# from django.shortcuts import render

# from django.contrib.auth.models import User
# from rest_framework import viewsets
# from .serializers import UsersSerializer
# from .models import Users

# # Create your views here.
# class UsersViewSet(viewsets.ModelViewSet):
#     """
#     API endpoint that allows users to be viewed or edited.
#     """
#     queryset = Users.objects.all()
#     serializer_class = UsersSerializer



from django.shortcuts import render 
from django.http import HttpResponse,FileResponse
from django.http.response import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from rest_framework.parsers import JSONParser 
from rest_framework import status

from django.http.response import StreamingHttpResponse as HttpResponse
     
from .models import Users,Questions,Categories,Selected_questions,Exam,Result_set,Email_status,Code_questions,Selected_code_questions,Coding_result_set,Admin_users,Email_open_status
from .serializers import UsersSerializer


from django.db import connection
import random
import json
import re,os
import datetime,time,pytz
# from dateutil import tz
from django.utils.dateparse import parse_datetime
from django.utils import timezone
from html import escape,unescape
from collections import OrderedDict 
from django.core.mail import EmailMultiAlternatives
import imaplib
import email
import threading 
from multiprocessing import Manager

     
# @csrf_exempt
# def user_signup(request):
#     if request.method == 'POST':
#         user_data = JSONParser().parse(request)
#         users = UsersSerializer(data=user_data)
#         if users.is_valid():
#             if(users.save()):
#             return JsonResponse(users.data, status=status.HTTP_201_CREATED) 
#         return JsonResponse(users.errors, status=status.HTTP_400_BAD_REQUEST)
  
 
from django.contrib.auth.models import User
import hashlib 
from Crypto.Cipher import AES
import base64
KEY="This is a key123"


IMAP_SERVER='imap.gmail.com'
EMAIL_HOST_USER = 'terralogic.smarthire@gmail.com'
EMAIL_HOST_PASSWORD = 'vjtohoaprxvfxfab'


def do_enc(email_id,password=None):    
    obj = AES.new(KEY, AES.MODE_CBC, 'This is an IV456')
    if password==None:
        password = User.objects.make_random_password(length=6) 
    sha_code=hashlib.sha256(email_id.strip().encode()).hexdigest()
    enc_pwd=obj.encrypt(password+(16-len(password))*"=")
    return sha_code[:10]+base64.b64encode(enc_pwd).decode()

def do_dec(password): 
    try:
        enc_pwd=base64.b64decode(password[10:].encode())
        obj = AES.new(KEY, AES.MODE_CBC, 'This is an IV456')
        return str(obj.decrypt(enc_pwd).decode("utf-8")).replace("=","")
    except Exception as e:
        print(str(e))
        return "Failed"

@csrf_exempt
def add_admin(request):
    if  request.method == 'POST' and ("admin" in request.session) and (request.session["admin"]!=None):

        if len(Admin_users.objects.filter(email=request.session["admin"],super_admin=True).values())==0:
            return HttpResponse("error&sep;You are not permitted to this operation",content_type="text")
        try:
            email=request.POST.get("email").strip().lower()
            admin=Admin_users()
            admin.name=request.POST.get("name").strip()
            admin.email=email
            admin.password=do_enc(email,request.POST.get("password").strip())
            admin.super_admin=request.POST.get("super_admin").strip()
            admin.save()
            if admin!=None:
                return HttpResponse("success&sep;Registered successfully",content_type="text")
            else:
                return HttpResponse("error&sep;Registration Failed",content_type="text")
        except Exception as e:
            print(str(e))
            return HttpResponse("error&sep;Failed",content_type="text")
    else:
        return HttpResponse("Only permitted to admin",content_type="text")

@csrf_exempt
def view_admin(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        print(len(Admin_users.objects.filter(email=request.session["admin"],super_admin=True).values()))
        if len(Admin_users.objects.filter(email=request.session["admin"],super_admin=True).values())!=0:
            orig_json=[dict(item) for item in Admin_users.objects.all().values('id','name','email','password','super_admin')]
            for i in range(0,len(orig_json)):
                if orig_json[i]["password"]!="":
                    orig_json[i]["password"]=do_dec(orig_json[i]["password"])
            return HttpResponse(json.dumps(orig_json),content_type="text")
        else:
            stmt='''
            SELECT 
                id,
                name,
                email,
                IF(email="'''+request.session["admin"]+'''",password,"") as "password",
                super_admin
            FROM
                smart_admin_users 
            '''      
            orig_json=json.loads(make_query(stmt))
            for i in range(0,len(orig_json)):
                if orig_json[i]["password"]!="":
                    orig_json[i]["password"]=do_dec(orig_json[i]["password"])
            return HttpResponse(json.dumps(orig_json),content_type="text")
    else:
        return HttpResponse("Only permitted to admin",content_type="text")


@csrf_exempt
def remove_admin(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        if len(Admin_users.objects.filter(email=request.session["admin"],super_admin=True).values())!=0:
            try:
                int_list=list(map(int,request.POST.get("ids").split(",")))
                total_admin=len(Admin_users.objects.filter(email=request.session["admin"],super_admin=True).values())
                admin_to_be_deleted=len(Admin_users.objects.filter(id__in=int_list,email=request.session["admin"],super_admin=True).values())
                if total_admin-admin_to_be_deleted==0:
                    return HttpResponse("error&sep;Not allowed to delete all the admins",content_type="text")
                Admin_users.objects.filter(id__in=int_list).delete()
                return HttpResponse("success&sep;User Deleted Successfully",content_type="text")
            except Exception as e:
                print(str(e))
                return HttpResponse("error&sep;Failed to delete user",content_type="text")
        else:
            return HttpResponse("error&sep;You are not permitted to this operation",content_type="text")
        
    else:
        return HttpResponse("Only permitted to admin",content_type="text")

@csrf_exempt
def update_admin(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        pwd=""
        try:
            test=Admin_users.objects.filter(email=request.session["admin"],super_admin=True).values()
            print(len(test),test)
            pwd=request.POST.get("password").strip()
            if len(pwd)<4 or len(pwd)>10:
                return HttpResponse("error&sep;Invalid password length",content_type="text")
        except Exception as e:
            print(str(e))
            return HttpResponse("error&sep;An error occurred",content_type="text")
        if len(Admin_users.objects.filter(email=request.session["admin"],super_admin=True).values())!=0:
            email=""
            try:
                email=Admin_users.objects.filter(id=request.POST.get("uid")).values()[0]["email"]
                if len(Admin_users.objects.filter(super_admin=True).values())==1 and int(request.POST.get("super_admin"))==0:
                    return HttpResponse("error&sep;There should be at least one admin - "+email,content_type="text")
                password=do_enc(email,pwd)
                Admin_users.objects.filter(id=request.POST.get("uid")).update(password=password,super_admin=request.POST.get("super_admin"))
                return HttpResponse("success&sep;"+str(email)+" : Updated successfully",content_type="text")
            except Exception as e:
                print(str(e))
                return HttpResponse("error&sep;Failed to update user "+email,content_type="text")
        else:
            try:
                email=request.session["admin"]
                password=do_enc(email,pwd)
                resp=Admin_users.objects.filter(email=email).update(password=password)
                if resp>0:
                    return HttpResponse("success&sep;"+str(email)+" : Updated successfully",content_type="text")
                else:
                    return HttpResponse("error&sep;Failed to update user "+email,content_type="text")
            except Exception as e:
                print(str(e))
                return HttpResponse("error&sep;Failed to update user "+email,content_type="text")
        
    else:
        return HttpResponse("Only permitted to admin",content_type="text")

@csrf_exempt
def user_signup(request):
    if  request.method == 'POST' and ("admin" in request.session) and (request.session["admin"]!=None):
        try:
            users = Users()
            users.name=request.POST.get("name").strip()
            email_id=request.POST.get("email").strip()
            regexp_str="^[A-Za-z0-9!#$%&'*+\/=?^_`{|}~-]+(?:\.[A-Za-z0-9!#$%&'*+\/=?^_`{|}~-]+)*@(?:[A-Za-z0-9](?:[A-Za-z0-9-]*[A-Za-z0-9])?\.)+[A-Za-z0-9](?:[A-Za-z0-9-]*[A-Za-z0-9])?$"
            if not re.match(regexp_str,email_id):
                return HttpResponse("error&sep;Invalid email id",content_type="text")
            users.email=email_id
            users.exam_id=request.POST.get("exam")
            users.password=do_enc(email_id)
            if not request.POST.get("mob").isnumeric():
                return HttpResponse("error&sep;Incorrect value for mobile number "+request.POST.get("mob"),content_type="text")
            if len(request.POST.get("mob"))!=10:
                return HttpResponse("error&sep;Length of mobile number is incorrect",content_type="text")
            users.mobile_no=request.POST.get("mob")
            if re.match("^(127\.0\.0\.1)|localhost",request.get_host()):
                users.password=do_enc(email_id,"passme")
            users.score=-1
            users.save()
            if users!=None:
                mythread=threading.Thread(target=delete_mail_delivery_status, args=(email_id,))
                mythread.start()
                # mythread.join()
                mythread=threading.Thread(target=delete_mail, args=(email_id,))
                mythread.start()
                # mythread.join()
                return HttpResponse("success&sep;Registered successfully",content_type="text")
            else:
                return HttpResponse("error&sep;Registration Failed",content_type="text")
        except Exception as e:
            print(str(e))
            return HttpResponse("error&sep;Failed",content_type="text")
    else:
        return HttpResponse("Only permitted to admin",content_type="text")


@csrf_exempt
def user_login(request):
    request.session["logged_in"]=None
    if request.method == 'POST' and 'email' in request.POST and 'password' in request.POST: 
        try:
            request.session.clear()
            email_id=request.POST.get("email").strip().lower()
            user_pwd=request.POST.get('password').strip()
            if len(user_pwd)<4 or len(user_pwd)>10:
                return HttpResponse("Password should consist of minimum 4 character and maximum 10 character, current length is "+str(len(user_pwd)),content_type="text")
            regexp_str="^[A-Za-z0-9!#$%&'*+\/=?^_`{|}~-]+(?:\.[A-Za-z0-9!#$%&'*+\/=?^_`{|}~-]+)*@(?:[A-Za-z0-9](?:[A-Za-z0-9-]*[A-Za-z0-9])?\.)+[A-Za-z0-9](?:[A-Za-z0-9-]*[A-Za-z0-9])?$"
            if not re.match(regexp_str,email_id):
                return HttpResponse("Invalid email id",content_type="text")
            try:
                pwd=Admin_users.objects.filter(email=email_id).values()[0]["password"]
                if pwd==do_enc(email_id,user_pwd):
                    request.session["admin"]=email_id
                    return HttpResponse("admin",content_type="text")
                else:
                    return HttpResponse("Invalid admin credentials",content_type="text")
            except Exception as e:
                pass              

            try:
                uid=Users.objects.filter(email=email_id).values()[0]["id"]
                if int(json.loads(get_all_exam_status(uid))[0]["status_code"])==3:
                    return HttpResponse("You have already attended exam",content_type="text")
            except:
                return HttpResponse("Email id is not registered",content_type="text")
           
            validuser = Users.objects.filter(email=email_id,password__exact=do_enc(email_id,user_pwd))
            usersserializer = UsersSerializer(validuser, many=True)
            if(len(usersserializer.data)>0):
                request.session["logged_in"]=email_id
                return HttpResponse("valid",content_type="text")
            else:
                return HttpResponse("Invalid Credentials!",content_type="text")
        except Exception as e:
            print(str(e))
            return HttpResponse("An error occurred")
    else:
        return HttpResponse("Failed to login")


@csrf_exempt
def get_login_status(request):
    txt=""    
    # for key, val in request.session.items():
    #     txt+=key+"=>"+val+"\n";
    # txt+="\n\n"
    # auth = re.compile(r'^myAuth_.+$')
    # for header in request.META:
    #     if auth.match(header):
    #         return HttpResponse(request.META[header],content_type="text")
    #     txt+=header+"\t"+str(request.META[header])+"\n"
    if ("logged_in" in request.session) and (request.session["logged_in"]!=None):
        txt+="user in: "+request.session["logged_in"]
    else:
        txt+="No user logged in:"
    return HttpResponse(txt,content_type="text")
    # return JsonResponse(usersserializer.data, safe=False)


@csrf_exempt
def get_admin_status(request):
    txt=""    
    # for key, val in request.session.items():
    #     txt+=key+"=>"+val+"\n";
    # txt+="\n\n"
    # auth = re.compile(r'^myAuth_.+$')
    # for header in request.META:
    #     if auth.match(header):
    #         return HttpResponse(request.META[header],content_type="text")
    #     txt+=header+"\t"+str(request.META[header])+"\n"
    if ("admin" in request.session) and (request.session["admin"]!=None):
        txt+="user in: "+request.session["admin"]
    else:
        txt+="No user logged in:"
    return HttpResponse(txt,content_type="text")
    # return JsonResponse(usersserializer.data, safe=False)   


@csrf_exempt
def set_sess(request):
    request.session["guru"]=str(datetime.datetime.now())
    return HttpResponse(request.session["guru"],content_type="text")

    
@csrf_exempt
def del_sess(request):
    request.session.clear()
    return HttpResponse("done",content_type="text")

@csrf_exempt
def addcat(request):  
    if  request.method == 'POST' and ("admin" in request.session) and (request.session["admin"]!=None):
        try:
            if len(request.POST.get("cat"))<2 :
                raise Exception('values are null')
            print(request.POST.get("cat"))
            categories = Categories()
            categories.cat=request.POST.get("cat").strip()
            try:
                categories.save()
            except Exception as e:
                print(str(e))
                return HttpResponse("Category already exists",content_type="text")  
            if categories!=None:
                return HttpResponse("Category added",content_type="text")
            else:
                return HttpResponse("Failed",content_type="text")     
        except Exception as e:
            print(str(e))
            return HttpResponse("Failed",content_type="text")   
    else:
        return HttpResponse("invalid req",content_type="text")  


@csrf_exempt
def add_exam(request):  
    if  request.method == 'POST' and ("admin" in request.session) and (request.session["admin"]!=None):
        try:
            today=datetime.datetime.today()
            s_date=datetime.datetime.strptime(request.POST.get("s_date"), "%d/%m/%Y %H:%M")
            e_date=datetime.datetime.strptime(request.POST.get("e_date"), "%d/%m/%Y %H:%M")
            print(today,s_date,e_date)
            if(today>s_date):
                return HttpResponse("Inavalid start time/date",content_type="text")  
            if(s_date>e_date):
                return HttpResponse("Inavalid time/date range",content_type="text")  
            if(len(request.POST.get("e_name"))<=1):
                return HttpResponse("Enter exam name",content_type="text")  
                        
            exam = Exam()
            exam.e_name=request.POST.get("e_name").strip()
            exam.start_date=s_date
            exam.end_date=e_date
           
            try:
                exam.save()
            except Exception as e:
                print(str(e))
                return HttpResponse("Exam already exists",content_type="text")  
            if exam!=None:
                return HttpResponse("success",content_type="text")
            else:
                return HttpResponse("Failed",content_type="text")                
        except Exception as e:
            print(str(e))
            return HttpResponse("Failed",content_type="text")   
    else:
        return HttpResponse("invalid req method",content_type="text")  

@csrf_exempt
def remaining_exam_duration(request): 
    if  request.method == 'POST' and ("admin" in request.session) and (request.session["admin"]!=None):
        try:
            stmt="SELECT abs(TIMESTAMPDIFF(SECOND,se.start_date,se.end_date))-(duration+code_duration) as remaining FROM smart_exam as se WHERE id="+str(request.POST.get("exam"))+";"
            remaining=json.loads(make_query(stmt))[0]
            print(remaining)
            return HttpResponse(json.dumps(remaining),content_type="text")  
        except Exception as e:
            print(str(e))
            return HttpResponse("warning&sep;Enter valid details",content_type="text")  
    else:
        return HttpResponse("error&sep;You are not allowed for do this operation",content_type="text")  

        


@csrf_exempt
def add_que_set(request):  
    if  request.method == 'POST' and ("admin" in request.session) and (request.session["admin"]!=None):
        try:
            print(request.POST.get("exam"))
            myjson=json.loads(request.POST.get("data"))
            print(str(myjson))
            total_mark=0
            try:
                stmt="SELECT abs(TIMESTAMPDIFF(SECOND,se.start_date,se.end_date)) as diff FROM smart_exam as se WHERE id="+str(request.POST.get("exam"))+";"
                max_dur=json.loads(make_query(stmt))[0]["diff"]
                dur=Exam.objects.filter(id=request.POST.get("exam")).values()[0]["code_duration"]
                print(int(dur)+abs(int(request.POST.get("dur")))*60,max_dur)
                if (int(dur)+abs(int(request.POST.get("dur")))*60)>int(max_dur):
                    return HttpResponse("error&sep;Maximum exam durartion is "+str(int(int(max_dur)/60))+" Mins",content_type="text")

            except:
                pass
            try:
                Selected_questions.objects.filter(exam_id=request.POST.get("exam")).delete()
            except Exception as e:
                print(str(e))
            for val in myjson:
                cursor = connection.cursor()     
                exam=Exam.objects.filter().values()[0]   
                stmt="select id from smart_questions where cat_id='"+str(val["cat"])+"' order by RAND() LIMIT "+val["total"]+";"
                cursor.execute(stmt)
                questions= cursor.fetchall() 
                #questions=Questions.objects.filter(cat=str(val["cat"])).values()[0:int(val["total"])]
                total_mark+=len(questions)
                for val2 in questions:
                    selected_questions=Selected_questions()
                    selected_questions.que_id=val2[0]
                    selected_questions.exam_id=request.POST.get("exam")
                    cursor = connection.cursor()
                    duration=abs(int(request.POST.get("dur")))*60
                    stmt="UPDATE smart_exam SET duration = \'"+str(duration)+"\',total_marks = \'"+str(total_mark)+"\' WHERE id=\'"+request.POST.get("exam")+"\';"
                    cursor.execute(stmt)
                    cursor.close()
                    try:
                        selected_questions.save()
                    except:
                        pass  

            return HttpResponse("success&sep;Exam Populated",content_type="text")  

        except Exception as e:
            return HttpResponse("warning&sep;Enter valid details",content_type="text")  
    else:
        return HttpResponse("Invalid req",content_type="text")  


@csrf_exempt
def addque(request):      
    if ("admin" in request.session) and (request.session["admin"]!=None):
        try:
            if len(request.POST.get("quetions"))<2 or len(request.POST.get("opt1"))<1 or len(request.POST.get("opt2"))<1 or len(request.POST.get("opt3"))<1 or len(request.POST.get("opt4"))<1 or len(request.POST.get("ans"))<1:
                raise Exception('values are null')
            questions = Questions()
            sent_que="<div>"+escape(request.POST.get("quetions")).replace("\n","<br />")+"</div>"
            if(request.POST.get("formated_para")!="nothing"):
                sent_que+="\n"+"<pre>"+escape(request.POST.get("formated_para"))+"</pre>"
            questions.question=sent_que
            questions.opt1=request.POST.get("opt1")
            questions.opt2=request.POST.get("opt2")
            questions.opt3=request.POST.get("opt3")
            questions.opt4=request.POST.get("opt4")
            questions.ans=request.POST.get("ans")
            questions.cat_id=request.POST.get("cat")
            questions.save()
            if questions!=None:
                return HttpResponse("success",content_type="text")
            else:
                return HttpResponse("Failed",content_type="text")     
        except Exception as e:
            print(str(e))
            return HttpResponse("Failed",content_type="text")   
    else:
        return HttpResponse("invalid req",content_type="text")  


@csrf_exempt
def set_cur_exam(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        try:
            cursor = connection.cursor()
            stmt="UPDATE smart_exam SET on_going_exam = \'false\';"
            cursor.execute(stmt)
            stmt="UPDATE smart_exam SET on_going_exam = \'true\' where id=\'"+request.POST.get("exam_id")+"\';"
            print(stmt+"\n")
            cursor.execute(stmt)
            print("Update:"+str(cursor.rowcount)+"\n")
            connection.commit()
            cursor.close()
            if(cursor.rowcount>0):
                return HttpResponse("sucess",content_type="text")  
            else:                
                return HttpResponse("no change",content_type="text")  
        except Exception as e:
            print(str(e))
            return HttpResponse("Failed",content_type="text")   
    else:
        return HttpResponse("invalid req",content_type="text")  

@csrf_exempt
def get_cat(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        stmt='''
        SELECT cat,c.id,count(q.cat_id) AS available
                FROM    smart_questions q
                RIGHT JOIN smart_categories c
                ON c.id=q.cat_id
                GROUP BY c.id
        '''
      
        return HttpResponse(make_query(stmt),content_type="text")
        # return HttpResponse(json.dumps([dict(item) for item in Categories.objects.all().values('cat','id')]),content_type="text")
    else:
        return HttpResponse("Invalid req",content_type="text")
@csrf_exempt
def get_exam(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        return HttpResponse(json.dumps([dict(item) for item in Exam.objects.all().values('e_name','id')]),content_type="text")
    else:
        return HttpResponse("Invalid req",content_type="text")

@csrf_exempt
def get_select_que_count(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        return HttpResponse(str(len(Selected_questions.objects.filter(exam_id=request.POST.get("exam_id")).values())),content_type="text")
    else:
        return HttpResponse("Invalid req",content_type="text")

@csrf_exempt
def get_questions(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        if(request.POST.get("id")=="all"):
            # return HttpResponse(json.dumps([dict(item) for item in Questions.objects.all().values('question','id','opt1','opt2','opt3','opt4','ans')]),content_type="text")
            stmt='''
            SELECT 
                    DISTINCT sq.id,question, opt1, opt2, opt3, opt4, ans,cat
                    FROM   smart_questions AS sq 
                    INNER JOIN smart_categories AS ct
                            ON sq.cat_id = ct.id
            '''
          
            return HttpResponse(make_query(stmt),content_type="text")
        elif(('type' in request.POST) and request.POST.get("type")=="exam"):
            stmt='''
            SELECT 
                     DISTINCT sq.id,question, opt1, opt2, opt3, opt4, ans,cat
                    FROM   smart_questions AS sq 
                    INNER JOIN smart_selected_questions AS ssq
                            ON ssq.que_id = sq.id
                    INNER JOIN smart_categories AS ct
                            ON sq.cat_id = ct.id
                    WHERE ssq.exam_id=\''''+request.POST.get("id")+'''\'
            '''

          
            return HttpResponse(make_query(stmt),content_type="text")
        else:
            stmt='''
            SELECT 
                     DISTINCT sq.id,question, opt1, opt2, opt3, opt4, ans,cat
                    FROM   smart_questions AS sq 
                    INNER JOIN smart_categories AS ct
                            ON sq.cat_id = ct.id
                    WHERE ct.id=\''''+request.POST.get("id")+'''\'
            '''
          
            return HttpResponse(make_query(stmt),content_type="text")
            # return HttpResponse(json.dumps([dict(item) for item in Questions.objects.filter(cat_id=request.POST.get("id")).values( 'question','id','opt1','opt2','opt3','opt4','ans')]),content_type="text")
    else:
        return HttpResponse("Invalid req",content_type="text")


# @csrf_exempt
# def get_sel_que(request):
#     if ("admin" in request.session) and (request.session["admin"]!=None):
#         if(request.POST.get("id")=="all"):
#             # return HttpResponse(json.dumps([dict(item) for item in Questions.objects.all().values('question','id','opt1','opt2','opt3','opt4','ans')]),content_type="text")
#             stmt='''
#             SELECT 
#                     DISTINCT sq.id,question, opt1, opt2, opt3, opt4, ans,cat
#                     FROM   smart_questions AS sq 
#                     INNER JOIN smart_categories AS ct
#                             ON sq.cat_id = ct.id
#             '''
#           
#             return HttpResponse(make_query(stmt),content_type="text")
#         elif(('type' in request.POST) and request.POST.get("type")=="exam"):

# def check_attend(request):  
#     user = Users.objects.filter(email=request.session["logged_in"]).values()
#     return len(user.get("score"))>0

def check_attend(request):  
    user = Users.objects.filter(email=request.session["logged_in"],score="-1").values()
    return len(user)>0


def time_to_sec(cur_time):
    cur_time_arr=cur_time.split(":")
    print(cur_time_arr)
    cur_time_in_sec=int(cur_time_arr[2])+int(cur_time_arr[1])*60+int(cur_time_arr[0])*3600
    return cur_time_in_sec

from django.db.models import Q

@csrf_exempt
def get_q(request):  
    #expire session on browser close
    request.session.set_expiry(0)   
    print("\n\n--------------------------------------------------------------------")
    if request.method == 'POST' and (check_attend(request) or "qno" in request.session):
        print("----------------get_q----------------------------------------------------")
        if ("qno" not in request.session):
            request.session["qno"]=0
        else:
            request.session["qno"]+=1
        q_str=""
        txt=""
        q_set=[]

        today=timezone.now()

        try:
            exam=Exam.objects.filter(id=(Users.objects.filter(email=request.session["logged_in"]).values())[0]["exam_id"],start_date__lte=today).values()[0]
        except Exception as e:
            print("range: "+str(e))
            return HttpResponse("&start;",content_type="text")
        try:
            exam=Exam.objects.filter(id=(Users.objects.filter(email=request.session["logged_in"]).values())[0]["exam_id"],end_date__gte=today).values()[0]
        except Exception as e:
            print("range: "+str(e))
            return HttpResponse("&close;",content_type="text")
        if ("timer" not in request.session):
            request.session["timer"]=datetime.datetime.now().strftime("%H:%M:%S")
        
        if(time_to_sec(datetime.datetime.now().strftime("%H:%M:%S"))-time_to_sec(request.session["timer"])>exam["duration"]):
            return HttpResponse("&timeout;",content_type="text")
        #set initial score=0 to mark as attended
        
        # print(quetions)     
        #save prepared/shuffled quetion to session

       

        if("quetion_set" not in request.session):
            cursor = connection.cursor()        
            stmt="select q.id,q.question,q.opt1,q.opt2,q.opt3,q.opt4,q.ans,sq.exam_id from smart_selected_questions as sq INNER JOIN smart_questions as q ON  sq.que_id=q.id where sq.exam_id=\'"+str(exam["id"])+"\';"
            cursor.execute(stmt)
            quetions= cursor.fetchall()   
            cursor = connection.cursor()         
            if(len(quetions)<1):
                return HttpResponse("&empty;",content_type="text")
            stmt="UPDATE smart_users SET score = \'0\' WHERE email=\'"+request.session["logged_in"]+"\' and score = \'-1\';"
            cursor.execute(stmt)
            print(stmt+"\nUpdate smart_users:"+str(cursor.rowcount)+"\n")
            cursor.close()
            
            for val in quetions:
                ind_que=list()
                for unescape_col in val:
                    print(unescape_col)
                    ind_que.append(unescape(str(unescape_col)))
                ind_que[1]=str(val[1])
                txt=str(ind_que[0])+"&sep;"+("&sep;".join(ind_que[1:len(ind_que)-1]))
                txt+="&sep;"+str(ind_que[len(ind_que)-1])
                q_str+=txt+"&diff"
            q_set=q_str.split("&diff")
            print("\n prepared quetion saved to session\n")
            request.session["quetion_set"]=q_set[:len(q_set)-1]
            random.shuffle(request.session["quetion_set"])
            print(str(request.session["quetion_set"])+"\n")
        # print("\n"+str(request.session["quetion_set"]))
        if(len(request.session["quetion_set"])>request.session["qno"]):
            quetion_to_be_send=request.session["quetion_set"][request.session["qno"]]
            quetion_to_be_send_arr=quetion_to_be_send.split("&sep;")
            exam_id=quetion_to_be_send_arr[len(quetion_to_be_send_arr)-1]
            #remove answer
            quetion_to_be_send_arr=quetion_to_be_send_arr[:len(quetion_to_be_send_arr)-2]
            quetion_to_be_send="&sep;".join(quetion_to_be_send_arr)
            print(quetion_to_be_send+"&sep;"+str(request.session["qno"]+1)+"&sep;"+str(len(request.session["quetion_set"])))
            
            result_set=Result_set()
            result_set.que_id=quetion_to_be_send_arr[0]
            result_set.user_id=Users.objects.filter(email=request.session["logged_in"]).values()[0]["id"]
            result_set.exam_id=exam_id
            result_set.date_f=datetime.date.today().strftime("%Y-%m-%d")
            result_set.s_time=datetime.datetime.now().strftime("%H:%M:%S")
            result_set.e_time=datetime.datetime.now().strftime("%H:%M:%S")
            result_set.save()
            if result_set!=None:
                print("\nsucess:")
            else:
                print("\nFailed: ")
            # except Exception as e:
            #     print("\ninsert: "+str(e))
            return HttpResponse(quetion_to_be_send+"&sep;"+str(request.session["qno"]+1)+"&sep;"+str(len(request.session["quetion_set"]))+"&sep;"+str(exam["duration"]),content_type="text")
        else:
            return HttpResponse("&end;",content_type="text")
    else:
        return HttpResponse("closed")



from django.db import connection
@csrf_exempt
def ver_q(request):  
    if ("logged_in" in request.session) and (request.session["logged_in"]!=None):
        print("----------------ver_q----------------------------------------------------")
        qid=request.POST.get('id')
        ans=request.POST.get('ans')

        if ("prev_que" not in request.session) or (request.session["prev_que"]==None):
            request.session["prev_que"]=qid
        elif(request.session["prev_que"]==qid):
            print("no inc")
            return HttpResponse("no inc",content_type="text")

        request.session["prev_que"]=qid
        print("Recived from candidate")
        print(str(request.session["qno"])+")"+request.session["logged_in"]+","+str(qid)+","+ans+"\n")

        # print(str(request.session["quetion_set"])+"\n")

        quetion_sent=request.session["quetion_set"][request.session["qno"]]
        quetion_sent_arr=quetion_sent.split("&sep;")
        exam_id=quetion_sent_arr[len(quetion_sent_arr)-1]
        #from stored values in session
        #print(str(quetion_sent_arr)+"\n")
        print("In the DB")
        print("id="+quetion_sent_arr[0]+","+"ans="+quetion_sent_arr[len(quetion_sent_arr)-2]+"\n")

        if ("score" not in request.session):
            request.session["score"]=0   
        try:
            cursor = connection.cursor()
            stmt="UPDATE smart_result_set SET ans = \'"+escape(str(ans))+"\',e_time=\'"+datetime.datetime.now().strftime("%H:%M:%S")+"\' WHERE user_id=\'"+str(Users.objects.filter(email=request.session["logged_in"]).values()[0]["id"])+"\' and  que_id=\'"+str(qid)+"\' and exam_id=\'"+exam_id+"\' ;"
          
            cursor.execute(stmt)
            print("UPDATE smart_result_set: "+str(cursor.rowcount)+"\n")
        except Exception as e:
            print(str(e))

        print("----Current score Before Match: "+str(request.session["score"]))
        if(re.match("^"+re.escape(str(ans))+"$",quetion_sent_arr[len(quetion_sent_arr)-2])):
            request.session["score"]+=1
            print("----Current score After Match:  "+str(request.session["score"]))
            try:
                cursor = connection.cursor()
                exam=Exam.objects.filter(id=(Users.objects.filter(email=request.session["logged_in"]).values())[0]["exam_id"]).values()[0]
                print("---------Exam: "+str(exam))
                stmt="UPDATE smart_users SET score = \'"+str(request.session["score"])+"/"+str(exam["total_marks"])+"\' WHERE email=\'"+request.session["logged_in"]+"\';"
                print(stmt+"\n")
                cursor.execute(stmt)
                print("Update smart_users:"+str(cursor.rowcount)+"\n")
                connection.commit()
                cursor.close()
            except Exception as e:
                print("\n upadte:"+str(e))
            return HttpResponse("inc",content_type="text")
        else:        
            return HttpResponse("no inc",content_type="text")
    else:
            return HttpResponse("Invalid req",content_type="text")


@csrf_exempt
def feedback(request):  
    if("logged_in" in request.session) and (request.session["logged_in"]!=None):
        try:
            cursor = connection.cursor()
            stmt="UPDATE smart_users SET feedback="+str(request.POST.get("feedback"))+" WHERE email=\'"+str(request.session["logged_in"])+"\';"
          
            cursor.execute(stmt)
            print("UPDATE smart_result_set: "+str(cursor.rowcount)+"\n")
            return HttpResponse("update status: "+str(cursor.rowcount),content_type="text")
        except Exception as e:            
            return HttpResponse("update failed",content_type="text")
    else:
        return HttpResponse("Only admin can",content_type="text")

 

from django.db import connection
@csrf_exempt
def count_malpractices(request):  
    if("logged_in" in request.session) and (request.session["logged_in"]!=None):
        try:
            cursor = connection.cursor()
            stmt="UPDATE smart_users SET malpractices=malpractices+1  WHERE email=\'"+str(request.session["logged_in"])+"\';"
          
            cursor.execute(stmt)
            print("UPDATE smart_result_set: "+str(cursor.rowcount)+"\n")
            return HttpResponse("update status: "+str(cursor.rowcount),content_type="text")
        except Exception as e:            
            return HttpResponse("update failed",content_type="text")
    else:
        return HttpResponse("Only admin can",content_type="text")
@csrf_exempt
def exam_status(request): 
    if check_attend(request):
        return HttpResponse("open",content_type="text")
    else:
        return HttpResponse("closed",content_type="text")

@csrf_exempt
def exam_logout(request): 
    if("logged_in" in request.session):
        del request.session['qno']
        return HttpResponse("Deleted",content_type="text")
    else:
        return HttpResponse("closed",content_type="text")



import xlwt

hex_no=8
def set_xl_color(workbook,color_name,r,g,b,bold):
    global hex_no
    font = xlwt.Font()
    font.name = 'Times New Roman'
    font.bold = bold
    if hex_no >=64:
        hex_no=8
    xlwt.add_palette_colour(color_name, hex_no) 
    workbook.set_colour_RGB(hex_no,r,g,b) 
    font.colour_index = xlwt.Style.colour_map[color_name]
    style = xlwt.XFStyle() 
    style.font = font
    hex_no+=1
    return style




@csrf_exempt
def view_res(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        max=";"
        if int(request.POST.get("max"))!=0:
            max="LIMIT "+request.POST.get("max")+";"
        if ('exam' in request.POST):
            stmt='''
                SELECT t1.ID,Username,Score,coding_score AS Score2,IF(total_dur1,total_dur1,0)+IF(total_dur2,total_dur2,0) AS 'Total Duration',Feedback,mobile_no AS "Mobile No" from
                    (
                        SELECT 
                            su.id AS ID,
                            su.email AS Username,
                            su.score AS Score,      
                            SUM(IF(q.ans=rs.ans,1,0)) AS score1,
                            SUM(IF(Round(Abs(TIME_TO_SEC(rs.s_time) - TIME_TO_SEC(rs.e_time)), 0),Round(Abs(TIME_TO_SEC(rs.s_time) -TIME_TO_SEC(rs.e_time)), 0),0)) AS total_dur1,
                            su.coding_score AS coding_score,
                            su.feedback AS Feedback,
                            su.mobile_no
                
                        FROM smart_result_set AS rs               
                            INNER JOIN smart_questions AS q
                                    ON q.id = rs.que_id 
                            RIGHT JOIN   smart_users AS su
                                    ON su.id = rs.user_id
                            WHERE su.exam_id=\''''+request.POST.get("exam")+'''\'
                            GROUP BY  su.id
                    ) as t1
                    LEFT JOIN
                    (
                        SELECT 
                            su.id AS ID  ,
                            count(*) AS count,
                            SUM(IF(Round(Abs(TIME_TO_SEC(crs.s_time) - TIME_TO_SEC(crs.e_time)), 0),Round(Abs(TIME_TO_SEC(crs.s_time) - TIME_TO_SEC(crs.e_time)), 0),0)) AS total_dur2,
                            SUM(crs.total_testcase_passed) AS score2
    
                        FROM smart_coding_result_set AS crs              
                        INNER JOIN   smart_users AS su
                                ON su.id = crs.user_id
                        WHERE su.exam_id= \''''+request.POST.get("exam")+'''\'
                        GROUP BY  su.id) as t2
                    ON t1.id=t2.id
                    ORDER BY IF(score1 is NULL,0,score1)+IF(score2 is NULL,0,score2) DESC,IF(total_dur1,total_dur1,0)+IF(total_dur2,total_dur2,0)          
            '''+max
            print(stmt)
            resp=make_query(stmt)
            json_data=json.loads(resp)
            workbook = xlwt.Workbook(encoding = 'ascii')
            worksheet = workbook.add_sheet('Smarthire')

            heading=set_xl_color(workbook,"head_color", 245, 66, 152,True)
            body=set_xl_color(workbook,"body_color",  27, 11, 77,False)
            inch=3333 # 3333 = 1" (one inch).



            worksheet.write(0, 0, label = 'User ID', style=heading)
            worksheet.write(0, 1, label = 'Email', style=heading)
            worksheet.write(0, 2, label = 'Score', style=heading)
            worksheet.write(0, 3, label = 'Coding score', style=heading)
            worksheet.write(0, 4, label = 'Total Duration(Mins)', style=heading)
            worksheet.write(0, 5, label = 'Feedback', style=heading)
            worksheet.write(0, 6, label = 'Mobile number', style=heading)


            for i in range(0,len(json_data)):
                score1=json_data[i]["Score"]
                score2=json_data[i]["Score2"]
                if score1=="-1":
                    score1="NA"
                if score2=="-1":
                    score2="NA"
                dur=0
                if json_data[i]["Total Duration"]!="None":
                    print(json_data[i]["Total Duration"])
                    dur=round(float(json_data[i]["Total Duration"])/60)
                worksheet.write(i+1, 0, label = json_data[i]["ID"], style=body)
                worksheet.write(i+1, 1, label = json_data[i]["Username"], style=body)
                worksheet.write(i+1, 2, label =score1, style=body)
                worksheet.write(i+1, 3, label = score2, style=body)
                worksheet.write(i+1, 4, label = str(dur), style=body)
                worksheet.write(i+1, 5, label = json_data[i]["Feedback"], style=body)
                worksheet.write(i+1, 6, label = json_data[i]["Mobile No"], style=body)

            for c in range(0,7):
                worksheet.col(c).width = round(inch*1.5)
            worksheet.col(1).width = round(inch*3)
            path=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))+"/../"
            print(path)
            workbook.save(path+'./docs/details.xls')
            workbook.save(path+'./public/details.xls')
            os.system("sudo chmod 777  "+path+"./docs/details.xls "+path+"./public/details.xls")

            return HttpResponse(resp,content_type="text")
        else:        
            stmt='''
             SELECT t1.ID,Username,Score,coding_score AS Score2,total_dur AS 'Total Duration',Feedback from
                    (
                        SELECT 
                            su.id AS ID,
                            su.email AS Username,
                            su.score AS Score,      
                            SUM(IF(q.ans=rs.ans,1,0)) AS score1,
                            SUM(Round(Abs(TIMEDIFF(rs.s_time , rs.e_time)), 0)) AS total_dur,
                            su.coding_score AS coding_score,
                            su.feedback AS Feedback
                
                        FROM smart_result_set AS rs               
                            INNER JOIN smart_questions AS q
                                    ON q.id = rs.que_id 
                            RIGHT JOIN   smart_users AS su
                                    ON su.id = rs.user_id
                            GROUP BY  su.id
                    ) as t1
                    LEFT JOIN
                    (
                        SELECT 
                            su.id AS ID  ,
                            count(*) AS count,
                            SUM(crs.total_testcase_passed) AS score2
    
                        FROM smart_coding_result_set AS crs              
                        INNER JOIN   smart_users AS su
                                ON su.id = crs.user_id
                        GROUP BY  su.id) as t2
                    ON t1.id=t2.id
                    ORDER BY score1+IF(score2 is NULL,0,score2) DESC,total_dur   
            '''+max
          
            return HttpResponse(make_query(stmt),content_type="text")
    else:
        return HttpResponse("Invalid req",content_type="text")

from decimal import Decimal
def default(o):
    if isinstance(o, (datetime.date, datetime.datetime)):
        return o.isoformat()
    if isinstance(o, (datetime.time, datetime.datetime)):
        return o.isoformat()
    if isinstance(o, Decimal):
        return str(o)

# def get_stmt(myregex,fields_dict,dict_arr):
#     temp_arr=[None]*len(dict_arr)
#     temp_str=[None]*len(dict_arr)
#     for val in fields_dict:
#         if(re.search(myregex,val,re.IGNORECASE)):
#             for dict_arr_val in dict_arr:
#                 temp_arr[dict_arr_val].append(dict_arr[dict_arr_val][val])
#     if len(temp_arr[0])>0:
#         for dict_arr_val in dict_arr:
#             temp_str[dict_arr_val]=",".join(temp_arr[dict_arr_val])
#     return temp_str



def utc_to_ist(local):
    # METHOD 1: Hardcode zones:
    from_zone = pytz.timezone('UTC')
    to_zone = pytz.timezone('Asia/Kolkata')

    # # METHOD 2: Auto-detect zones:
    # from_zone = tz.tzutc()
    # to_zone = tz.tzlocal()

    # utc = datetime.utcnow()
    utc = datetime.datetime.strptime(local, '%Y-%m-%d %H:%M:%S')

    # Tell the datetime object that it's in UTC time zone since 
    # datetime objects are 'naive' by default
    utc = utc.replace(tzinfo=from_zone)

    # Convert time zone
    return utc.astimezone(to_zone).strftime("%Y-%m-%d %H:%M:%S")

def make_query(stmt):
    cursor = connection.cursor() 
    try:
        cursor.execute(stmt)
    except:
        return "syntax error"
    det=cursor.fetchall()
    # print(det)
    header=[]
    for val in cursor.description:
        header.append(val[0])
    header_arr=[]
    mylist=[]
    c=0
    for val in det:          
        try:
            c=0
            list1=OrderedDict()
            for value in val:
                orig_val=unescape(str(value))
                if re.search("^\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}(.\d{1,6})?$",orig_val):
                    orig_val=utc_to_ist(re.search("\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}", orig_val).group(0))
                list1.update(OrderedDict({cursor.description[c][0]: orig_val}))
                c+=1
            mylist.append(list1)
        except Exception as e:
            print(str(e))  
    return str(json.dumps(mylist,sort_keys=False,indent=1,default=default))

@csrf_exempt
def view_det_res(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):

        where_case=group_case=order_case=having_case=strict=cond=agg=dist=""
        user_id=request.POST.get("id")
        try:
            dist=request.POST.get("dist")[1:len(request.POST.get("dist"))]
        except:
            pass
        try:
            where_case=request.POST.get("where")[1:len(request.POST.get("where"))]
        except:
            pass
        try:
           group_case=request.POST.get("group")[1:len(request.POST.get("group"))]
        except:
            pass
        try:
           having_case=request.POST.get("having")[1:len(request.POST.get("having"))]
        except:
            pass
        try:
           agg=request.POST.get("agg")[1:len(request.POST.get("agg"))]
        except:
            pass
        try:
            order_case=request.POST.get("order")[1:len(request.POST.get("order"))]
        except:
            pass
        try:
           strict=request.POST.get("strict")[1:len(request.POST.get("strict"))]
        except:
            pass
        
        # having_case=request.POST.get("having")
        
        print("------------\n"+user_id+"\n"+where_case+"\n"+group_case+"\n"+having_case+"\n"+order_case+"\n"+strict)

        sql_var="""
        DECLARE @cnt INT(11),
        DECLARE @txt LONGTEXT
        
        
        """
        fields=""  
        where_con=""
        groups_by=""
        having=""
        orders_by=""
        limits=""

        group_arr=[]
        fields_arr=[]
        order_arr=[]
        where_arr=[]
        having_arr=[]


        fields_dict=dict()
        fields_dict["question"]="q.question"
        fields_dict["correct"]="q.ans AS 'Correct Answers'"
        fields_dict["submitted"]="IF(rs.ans='undefined' or rs.ans='null','-',rs.ans) AS 'Submitted'"
        fields_dict["duration"]=""" ABS(TIMEDIFF(rs.s_time , rs.e_time)) AS 'Duration'"""
        fields_dict["valid"]="IF(q.ans=rs.ans,1,0) AS 'Result'"

        fields_dict["totalduration"]="""SUM(IF(rs.s_time - rs.e_time < rs.s_time,
                Round(Abs(rs.s_time - rs.e_time), 0), '-')) AS 'Total Duration'"""
        fields_dict["totalvalid"]="SUM(IF(q.ans=rs.ans,1,0)) AS 'Total Score'"
        fields_dict["date"]="rs.date_f AS 'Date'"
        fields_dict["category"]="ct.cat AS 'Category'"
        fields_dict["score"]="IF(su.score >= 0, su.score, '-') AS score"
        fields_dict["id"]="rs.user_id AS 'User ID'"
        fields_dict["username"]="user.email AS 'Username'"
        fields_dict["email"]="user.email AS 'Email'"
        fields_dict["exid"]="s_exam.id AS 'Exam ID'"

        field_key=dict()
        field_key["question"]="q.question"
        field_key["correct"]="q.ans"
        field_key["submitted"]="rs.ans"
        field_key["duration"]="rs.s_time,rs.e_time"
        field_key["valid"]="q.ans,rs.ans"
        field_key["date"]="rs.date_f"
        field_key["category"]="ct.cat"
        field_key["score"]="su.score"
        field_key["id"]="rs.user_id"
        field_key["username"]="user.email"
        field_key["email"]="user.email"
        field_key["exid"]="s_exam.id"


        having_dict=dict()
        having_dict["duration"]="""ABS(TIMEDIFF(rs.s_time , rs.e_time))"""
        having_dict["valid"]="IF(q.ans=rs.ans,1,0)"
        having_dict["question"]="q.question"
        having_dict["correct"]="q.ans"
        having_dict["submitted"]="rs.ans"
        having_dict["date"]="rs.date_f"
        having_dict["category"]="ct.cat"
        having_dict["score"]="su.score"
        having_dict["id"]="rs.user_id"
        having_dict["username"]="user.email"
        having_dict["email"]="user.email"
        having_dict["exid"]="s_exam.id"

        agg_dict=dict()
        agg_dict["count"]="COUNT"
        agg_dict["sum"]="SUM"
        agg_dict["max"]="MAX"
        agg_dict["min"]="MIN"
        agg_dict["avg"]="AVG"


  


        if(user_id=="all"):
            fields=fields_dict["username"]+","
        else:
            where_con="WHERE "+field_key["id"]+"='"+user_id+"'"
        fields+="""
            q.question AS Questions,
            q.ans AS  'Correct Answer',
            IF(rs.ans='undefined' or rs.ans='null','-',rs.ans) AS 'Submitted Answer',
            IF(q.ans =rs.ans,'Correct','Wrong') AS 'Result',
            ABS(TIMEDIFF(rs.s_time , rs.e_time)) AS 'Duration(sec)',
            ct.cat AS Category
        """  
        dist_arr=dist.split(" ")
        if group_case!="all" and group_case.strip()!="":
            group_arr=[]
            fields_arr=[]
            myregex="("+group_case.strip().replace(" ","|")+")"
            for val in field_key:
                if(re.search("^"+myregex,val,re.IGNORECASE)):
                    flag=False
                    if(len(dist)>0):
                        for dist_val in dist_arr:
                            if(re.search(dist_val,val,re.IGNORECASE)):
                                flag=True
                                break
                    if flag==True:
                        fields_arr.append("DISTINCT "+(fields_dict[val]))
                    else:
                        fields_arr.append((fields_dict[val]))
                    group_arr.append(field_key[val])
            if len(fields_arr)>0:
                fields=fields_dict["totalvalid"]+","+fields_dict["totalduration"]+","+(",".join(fields_arr))
                groups_by="GROUP BY "+(",".join(group_arr))
        if having_case.strip()!="":
            having_arr=[]
            myregex_arr=having_case.strip().split(",")
            for val in having_dict:
                try:
                    if isinstance(having_arr, list):
                        for myregex in myregex_arr:
                            having_key_val= re.split("[=<>]{1,2}",myregex)
                            gap=len(myregex)-(len(having_key_val[0])+len(having_key_val[1]))
                            symbols=myregex[len(having_key_val[0]):len(having_key_val[0])+gap]
                            print(str(having_key_val)+","+val)
                            if(re.match(having_key_val[0],val,re.IGNORECASE)):                           
                                try:
                                    having_arr.append(agg_dict[agg]+"("+fields_dict[val].split("AS")[0].strip()+")"+symbols+int(having_key_val[1]))
                                except:
                                    having_arr.append(agg_dict[agg]+"("+fields_dict[val].split("AS")[0].strip()+")"+symbols+"'"+having_key_val[1]+"'")
                                
                except Exception as e:
                    print(str(e))
            if len(having_arr)>0:
                having+=" HAVING "+(" and ".join(having_arr))
                print(where_con)
                
        if where_case.strip()!="":
            where_arr=[]
            myregex_arr=where_case.strip().split(",")
            for val in field_key:
                try:
                    if isinstance(myregex_arr, list):
                        for myregex in myregex_arr:
                            where_key_val= re.split("[=<>]{1,2}",myregex)
                            gap=len(myregex)-(len(where_key_val[0])+len(where_key_val[1]))
                            symbols=myregex[len(where_key_val[0]):len(where_key_val[0])+gap]
                            print(str(where_key_val)+","+val)
                            if(re.match(where_key_val[0],val,re.IGNORECASE)):
                                print("matched")
                                if(strict=="yes"):
                                    try:
                                        where_arr.append(fields_dict[val].split("AS")[0].strip()+symbols+str(int(where_key_val[1])))
                                    except Exception as e:
                                        print(str(e))
                                        where_arr.append(fields_dict[val].split("AS")[0].strip()+symbols+"'"+where_key_val[1]+"'")
                                else:
                                    where_arr.append(field_key[val]+" LIKE '%"+where_key_val[1]+"%'")
                except Exception as e:
                    print(str(e))
            if len(where_arr)>0:
                where_con="WHERE "+(" and ".join(where_arr))
                if(user_id!="all"):
                    where_con+=" and "+field_key["id"]+"='"+user_id+"'"
                print(where_con)
        if order_case.strip()!="":
            order_arr=[]
            myregex_arr=order_case.strip().split(",")
            for val in field_key:
                for myregex in myregex_arr:
                    order_val=myregex.split(" ")
                    if len(order_val)>1:
                        print(str(order_val)+","+val)
                        if(re.search(order_val[0],val,re.IGNORECASE)):
                            if(order_val[1]=="desc"):
                                order_arr.append(field_key[val]+" DESC")
                            else:                        
                                order_arr.append(field_key[val]+" ASC")
                    else:
                        if(re.search(order_val[0],val,re.IGNORECASE)):                      
                            order_arr.append(field_key[val]+" ASC")
            if len(order_arr)>0:
                orders_by="ORDER BY "+(",".join(order_arr))


        
            

        #################################################################################
        stmt="""
        SELECT 
        """+fields+"""
        FROM   smart_users AS su
            INNER JOIN smart_result_set AS rs
                    ON su.id = rs.user_id
            INNER JOIN smart_questions AS q
                    ON rs.que_id = q.id
            INNER JOIN smart_categories AS ct
                    ON ct.id = q.cat_id  
            INNER JOIN smart_users AS user
                    ON su.id = user.id    
            INNER JOIN smart_exam AS s_exam
                    ON s_exam.id = rs.exam_id      """
        stmt+=" "+where_con+" "
        stmt+=" "+groups_by+" "
        stmt+=" "+having+" "
        stmt+=" "+orders_by+" "
        stmt+=";"
        print("\n"+stmt+"\n")

        return HttpResponse(make_query(stmt),content_type="text")
    else:
        return HttpResponse("Invalid req",content_type="text")

'''
SELECT 
			DISTINCT su.name AS Name,
            su.email AS Username,
            su.score AS Score,
            rs.date_f AS Date,
            s_exam.e_name AS Exam
        
        FROM   smart_users AS su
            INNER JOIN smart_result_set AS rs
                    ON su.id = rs.user_id
            INNER JOIN smart_users AS user
                    ON su.id = user.id    
            INNER JOIN smart_exam AS s_exam
                    ON s_exam.id = rs.exam_id       WHERE rs.user_id='12'       ;
'''


@csrf_exempt
def view_cat_res(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        stmt='''
        SELECT 
            e_name AS Exam,
            total_marks AS 'Total Marks',
            start_date AS Date,
            CEILING(ROUND((duration /60))) AS 'Total Time(Mins)'
                
        FROM   
        smart_exam WHERE id=\''''+request.POST.get("eid")+'''\'
        '''
        data=make_query(stmt)
        try:
            return HttpResponse(data,content_type="text")
        except Exception as e:
            print(str(e))
    else:
        return HttpResponse("Invalid req",content_type="text")

@csrf_exempt
def view_graph(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        # stmt='''
        # SELECT 
        #     s_cat.cat as category,count(*) as count
            
        #     FROM   smart_users AS su
        #         INNER JOIN smart_result_set AS rs
        #                 ON su.id = rs.user_id
        #         INNER JOIN smart_users AS user
        #                 ON su.id = user.id    
        #         INNER JOIN smart_exam AS s_exam
        #                 ON s_exam.id = rs.exam_id  
        #         INNER JOIN smart_questions AS s_que
        #             ON s_que.id = rs.que_id
        #         INNER JOIN smart_categories AS s_cat
        #             ON s_cat.id = s_que.cat_id  
                        
        #                     WHERE su.id=\''''+request.POST.get("id")+'''\' and s_que.ans=rs.ans
        #     GROUP BY(s_que.cat_id )
        # '''
        user_id=request.POST.get("id")
        where_case=""
        strict=""
        try:
            where_case=request.POST.get("where")[1:len(request.POST.get("where"))]
        except:
            pass

        try:
           strict=request.POST.get("strict")[1:len(request.POST.get("strict"))]
        except:
            pass
        where_con=""
        print("where_case:\t"+where_case)

        fields_dict=dict()
        fields_dict["question"]="s_que.question"
        fields_dict["correct"]="s_que.ans AS 'Correct Answers'"
        fields_dict["submitted"]="IF(rs.ans='undefined' or rs.ans='null','-',rs.ans) AS 'Submitted'"
        fields_dict["duration"]=""" ABS(TIMEDIFF(rs.s_time , rs.e_time)) AS 'Duration'"""
        fields_dict["valid"]="IF(s_que.ans=rs.ans,1,0) AS 'Result'"

        fields_dict["totalduration"]="""SUM(IF(rs.s_time - rs.e_time < rs.s_time,
                Round(Abs(rs.s_time - rs.e_time), 0), '-')) AS 'Total Duration'"""
        fields_dict["totalvalid"]="SUM(IF(s_que.ans=rs.ans,1,0)) AS 'Total Score'"
        fields_dict["date"]="rs.date_f AS 'Date'"
        fields_dict["category"]="s_cat.cat AS 'Category'"
        fields_dict["score"]="IF(su.score >= 0, su.score, '-') AS score"
        fields_dict["id"]="rs.user_id AS 'User ID'"
        fields_dict["username"]="su.email AS 'Username'"
        fields_dict["email"]="su.email AS 'Email'"
        fields_dict["exid"]="s_exam.id AS 'Exam ID'"



        field_key=dict()
        field_key["question"]="s_que.question"
        field_key["correct"]="s_que.ans"
        field_key["submitted"]="rs.ans"
        field_key["duration"]="rs.s_time,rs.e_time"
        field_key["valid"]="s_que.ans,rs.ans"
        field_key["date"]="rs.date_f"
        field_key["category"]="s_cat.cat"
        field_key["score"]="su.score"
        field_key["id"]="rs.user_id"
        field_key["username"]="su.email"
        field_key["email"]="su.email"
        field_key["exid"]="s_exam.id"

        if(user_id!="all"):
            where_con="WHERE "+field_key["id"]+"='"+user_id+"'"
       
        if where_case.strip()!="":
            where_arr=[]
            myregex_arr=where_case.strip().split(",")
            for val in field_key:
                try:
                    if isinstance(myregex_arr, list):
                        for myregex in myregex_arr:
                            where_key_val= re.split("[=<>]{1,2}",myregex)
                            gap=len(myregex)-(len(where_key_val[0])+len(where_key_val[1]))
                            symbols=myregex[len(where_key_val[0]):len(where_key_val[0])+gap]
                            print(str(where_key_val)+","+val)
                            if(re.match(where_key_val[0],val,re.IGNORECASE)):
                                print("matched")
                                if(strict=="yes"):
                                    try:
                                        where_arr.append(fields_dict[val].split("AS")[0].strip()+symbols+str(int(where_key_val[1])))
                                    except Exception as e:
                                        print(str(e))
                                        where_arr.append(fields_dict[val].split("AS")[0].strip()+symbols+"'"+where_key_val[1]+"'")
                                else:
                                    where_arr.append(field_key[val]+" LIKE '%"+where_key_val[1]+"%'")
                except Exception as e:
                    print(str(e))
            if len(where_arr)>0:
                where_con="WHERE "+(" and ".join(where_arr))
                if(user_id!="all"):
                    where_con+=" and "+field_key["id"]+"='"+user_id+"'"
                print(where_con)
       
        stmt='''
        SELECT 
            s_cat.cat as category,
 			SUM(IF(s_que.ans=rs.ans,1,0)) AS 'correct',
            SUM(IF(s_que.ans<>rs.ans,1,0)) AS 'wrong' ,
            count(*) AS Total,
            SUM(ABS(TIMEDIFF(rs.s_time , rs.e_time))) AS duration
            
            FROM   smart_users AS su
                INNER JOIN smart_result_set AS rs
                        ON su.id = rs.user_id
                INNER JOIN smart_exam AS s_exam
                        ON s_exam.id = rs.exam_id  
                INNER JOIN smart_questions AS s_que
                    ON s_que.id = rs.que_id
                INNER JOIN smart_categories AS s_cat
                    ON s_cat.id = s_que.cat_id  

                           '''+where_con+'''
            GROUP BY(s_que.cat_id )
        '''
        print("stmt\n"+stmt)
        data=make_query(stmt)
        try:
            return HttpResponse(data,content_type="text")
        except Exception as e:
            print(str(e))
    else:
        return HttpResponse("Invalid req",content_type="text")


@csrf_exempt
def view_res_user(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        stmt='''
        SELECT 
                    DISTINCT su.name AS Name,
                    su.email AS Username,
                    rs.date_f AS Date,
                    s_exam.e_name AS Exam,
                    su.mobile_no AS "Mobile number"
                
                FROM   smart_users AS su
                    INNER JOIN smart_result_set AS rs
                            ON su.id = rs.user_id
                    INNER JOIN smart_users AS user
                            ON su.id = user.id    
                    INNER JOIN smart_exam AS s_exam
                            ON s_exam.id = rs.exam_id       WHERE rs.user_id=\''''+request.POST.get("id")+'''\'
        '''
        return HttpResponse(make_query(stmt),content_type="text")
    else:
        return HttpResponse("Invalid req",content_type="text")

@csrf_exempt
def del_user(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        resp=""
        try:
            int_list=list(map(int,request.POST.get("ids").split(",")))
            for id in int_list:
                cmd="sudo userdel $(grep -oi '^smarthire_user_"+str(id)+"_[0-9][0-9]*' /etc/passwd)"
                print(cmd)
                print("deluser: \t",subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT).stdout.read().decode())
            print(str(request.POST.get("ids").split(","))+"\n"+str(int_list))
            resp=str(Result_set.objects.filter( user_id__in=int_list ).delete())+"\n"
            resp+=str(Coding_result_set.objects.filter( user_id__in=int_list ).delete())+"\n"
            if re.match("^(127\.0\.0\.1)|localhost",request.get_host()):
                Users.objects.filter(id__in=int_list).update(score='-1')
                Users.objects.filter(id__in=int_list).update(coding_score='-1')
            if request.POST.get("all")=="true":
                resp+=str(Users.objects.filter(id__in=int_list).delete())
            print(resp)
            return HttpResponse("sucess: "+resp,content_type="text")
        except Exception as e:
            print(str(e))
            return HttpResponse("fail",content_type="text")
    else:
        return HttpResponse("Invalid req",content_type="text")


@csrf_exempt
def del_cat(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        int_list=list(map(int,request.POST.get("ids").split(",")))
        print(str(request.POST.get("ids").split(","))+"\n"+str(int_list))
        resp=""
        try:
            que_set=list(Questions.objects.filter(cat_id__in=int_list).values("id"));
            q_id=[val["id"] for val in que_set]
            res_set=Result_set.objects.filter( que_id__in=q_id).values("user_id")
            u_id=[val["user_id"] for val in res_set]
            resp+="uid: "+str(u_id)
            resp+="\nResult_set: "+str(Result_set.objects.filter( user_id__in=u_id).delete())
            resp+="\nCategories :"+str(Categories.objects.filter(id__in=int_list).delete())
            print(resp)
            return HttpResponse(resp,content_type="text")
        except Exception as e:
            print(str(e))
            return HttpResponse("fail",content_type="text")
    else:
        return HttpResponse("Invalid req",content_type="text")


@csrf_exempt
def exam_del(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        try:
            int_list=list(map(int,request.POST.get("ids").split(",")))
            status=str(Exam.objects.filter(id__in=int_list).delete())
            return HttpResponse("sucess: "+status,content_type="text")
        except Exception as e:
            print(str(e))
            return HttpResponse("fail",content_type="text")
    else:
        return HttpResponse("Invalid req",content_type="text")


@csrf_exempt
def del_questions(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        int_list=list(map(int,request.POST.get("ids").split(",")))
        print(str(request.POST.get("ids").split(","))+"\n"+str(int_list))
        resp=""
        try:
            if(request.POST.get("from")=="all"):
                print("--from all")
                resp=str(Questions.objects.filter(id__in=int_list ).delete())+"\n"
            else:
                print("--from selected que")
                resp=str(Selected_questions.objects.filter(que_id__in=int_list ).delete())+"\n"
            print(resp)
            return HttpResponse("sucess: "+resp,content_type="text")
        except Exception as e:
            print(str(e))
            return HttpResponse("fail",content_type="text")
    else:
        return HttpResponse("Invalid req",content_type="text")

# sent_emails=list()
manager=Manager()
# sent_emails=manager.list()
email_sending_status=False
email_delivery_status=list()
thread_stop=False


def add_mails(status):
    try:
        email_status_obj=Email_status()
        email_status_obj.status=status
        print(email_status_obj.save())
    except Exception as e:
        print(str(e))



def get_inbox_mails(mail_id,request): 
    global email_delivery_status,thread_stop
    try:
        # print("---Start---",threading.current_thread().ident)
        mail = imaplib.IMAP4_SSL(IMAP_SERVER,993)
        mail.login(EMAIL_HOST_USER,EMAIL_HOST_PASSWORD)
        # for i in mail.list()[1]:
        #     print(i)
        # mail.select('inbox')
        mail.select('"[Gmail]/All Mail"')
        #get sent mail first
        type2, data2 = mail.search(None, 'To', mail_id)
        #get its delivery status
        typ, data = mail.search(None, 'FROM', "mailer-daemon@googlemail.com", 'BODY',mail_id)
        # type, data = mail.search(None,  'All')
        # type, data = mail.search(None, 'SUBJECT', 'Smart hire from terralogic')

        id_list = list(data[0].split())
        total=len(id_list)           
        print(mail_id,"delivary total: ",total) 
        print(mail_id,"send_to total: ",len(list(data2[0].split())))
        if len(list(data2[0].split()))==0 or len(id_list)>0:
            Users.objects.filter(email=mail_id).update(email_sent_status=False)
        else:
            Users.objects.filter(email=mail_id).update(email_sent_status=True)             
        if len(id_list)==0:
            print("--------------------No mails--Thread aborted...")
            mail.close()
            mail.logout()
            return False
        for i in id_list:
            if thread_stop:
                print("------------------------------Thread aborted...")
                break
            typ, data = mail.fetch(str(i.decode()), '(RFC822)' )
            for response_part in data:
                if isinstance(response_part, tuple):
                    try:                        
                        msg = email.message_from_string(response_part[1].decode())
                        # txt+=(str(msg.__dict__))+"\n\n"
                        # if("terralogic.smarthire@gmail.com"==msg["To"]):
                        #     continue
                        txt=msg["X-Failed-Recipients"]+"\t"
                        txt+=msg[ 'Date']
                        # print(txt)
                        email_delivery_status.append(txt)
                    except Exception as e:
                        print(mail_id+"-"+str(e))
        mail.close()
        mail.logout()
        # print("---End---",threading.current_thread().ident)
    except Exception as e:
        print("get_inbox_mails-"+str(e))

def delete_mail_delivery_status(mail_id):
    try:
        mail = imaplib.IMAP4_SSL(IMAP_SERVER,993)
        mail.login(EMAIL_HOST_USER,EMAIL_HOST_PASSWORD)
        mail.select('"[Gmail]/All Mail"')
        typ, data = mail.search(None, 'FROM', "mailer-daemon@googlemail.com", 'BODY',mail_id)
        # type, data = mail.search(None, 'SUBJECT', 'Smart hire from terralogic')
        id_list = data[0].split()  
        total=len(id_list)
        print("del delivery total: ",total) 
        for i in id_list:
            typ, data = mail.fetch(str(i.decode()), '(RFC822)' )
            for response_part in data:
                if isinstance(response_part, tuple):
                    try:
                        msg = email.message_from_string(response_part[1].decode())
                        if msg["X-Failed-Recipients"]==None:
                            continue
                        print("deleting1... "+msg["X-Failed-Recipients"])
                        mail.store(i, '+FLAGS', '\\Deleted')
                    except Exception as e:
                        print("deleting1: "+str(e))
        print("Delete status1:  ",mail.expunge())
        mail.close()
        mail.logout()
    except Exception as e:
        print("delete_mail_delivery_status-"+str(e))


def delete_mail(mail_id):
    try:
        mail = imaplib.IMAP4_SSL(IMAP_SERVER,993)
        mail.login(EMAIL_HOST_USER,EMAIL_HOST_PASSWORD)
        mail.select('"[Gmail]/All Mail"')
        typ, data = mail.search(None, 'To', mail_id)
        # type, data = mail.search(None, 'SUBJECT', 'Smart hire from terralogic')
        id_list = data[0].split()  
        total=len(id_list)
        print("del all total: ",total) 
        for i in id_list:
            typ, data = mail.fetch(str(i.decode()), '(RFC822)' )
            for response_part in data:
                if isinstance(response_part, tuple):
                    try:
                        msg = email.message_from_string(response_part[1].decode())
                        mail.store(i, '+X-GM-LABELS', '\\Trash')
                        print("deleting2... ",msg["To"])
                    except Exception as e:
                        print("deleting2: "+str(e))
        print("Move to trash status2:  ",mail.expunge())
        mail.select('[Gmail]/Trash')  # select all trash
        mail.store("1:*", '+FLAGS', '\\Deleted')  #Flag all Trash as Deleted
        print("delete from trash status2:  ",mail.expunge())
        mail.close()
        mail.logout()
    except Exception as e:
        print("delete_mail-"+str(e))

threads=list()

def check_email_status(request):
    global thread_stop,threads
    email_delivery_status.clear()
    users=Users.objects.all().values()
    # if len(threads)>=len(users):
    #     thread_stop=True
    #     threads.clear()
    thread_stop=True
    time.sleep(1)
    thread_stop=False
    ind=0
    for thread in threads:
        if not thread.isAlive():
            threads.pop(ind)
        ind+=1
    for user in users:
        threads.append(threading.Thread(target=get_inbox_mails, args=(user['email'],request,)))
    for thread in threads:
        try:
            if not thread.isAlive():
                thread.start()
        except Exception as e:
            print("email status: "+str(e))
    # for thread in threads:
    #     thread.join()

@csrf_exempt
def retrive_email_status(request):
    global email_delivery_status
    if ("admin" in request.session) and (request.session["admin"]!=None):
        check_email_status(request)
        return HttpResponse("Welcome",content_type="text")
    else:
        return HttpResponse("Only admin can",content_type="text")

@csrf_exempt
def email_status_in_db(request):
    global email_delivery_status
    if ("admin" in request.session) and (request.session["admin"]!=None):
        if not 'quick' in request.POST:
            check_email_status(request)
        if 'exam' in request.POST and request.POST.get("exam")!=None and request.POST.get("exam")!="undefined":
            print("exam",request.POST.get("exam"))
            return HttpResponse(json.dumps([dict(item) for item in Users.objects.filter(email_sent_status=False,exam_id=request.POST.get("exam")).values('email','exam')]),content_type="text")
        return HttpResponse(json.dumps([dict(item) for item in Users.objects.filter(email_sent_status=False).values('email','exam')]),content_type="text")
    else:
        return HttpResponse("Only admin can",content_type="text")


@csrf_exempt
def send_cred(request):
    global email_sending_status
    if  request.method == 'POST':
        if ("admin" in request.session) and (request.session["admin"]!=None) and ('exam' in request.POST):
            if email_sending_status:
                add_mails("Already a user sending emails,wait until all email were sent")
                return HttpResponse("Already a user sending emails,wait until all email were sent",content_type="text")
            # # sent_emails.clear()
            # sent_emails[:]=[]
            add_mails("Sending email...")
            Email_status.objects.all().delete()
            email_sending_status=True
            try:
                
                if 'all' in request.POST:
                    Users.objects.filter(exam=request.POST.get("exam")).update(email_sent_status=False)
                    users=Users.objects.filter(exam=request.POST.get("exam")).values()
                    print("sending all...")
                else:
                    users=Users.objects.filter(exam=request.POST.get("exam"),email_sent_status=False).values()
                    print("sending failed...")

                for row in users:  
                    print(row) 
                    mythread=threading.Thread(target=delete_mail_delivery_status, args=(row["email"],))
                    mythread.start()
                    mythread.join()          
                    body="""
                        <h2>Your Credentials are:</h2>
                        <div style='font-size:15px;'>
                            <b>Email: </b><span style='color:red'>"""+row["email"]+"""</span><br />
                            <b>Password: </b><span style='color:red'><i>"""+do_dec(row["password"])+"""</i></span><br />
                            <a href='http://"""+request.get_host()+"""/#/?email="""+row["email"]+"""'>Click here to start Exam</a><br />
                        </div>
                        <div class="gmail_signature" data-smartmail="gmail_signature">
                        <div dir="ltr">
                            <br> 
                            <div>
                            <div dir="ltr">
                                <img src='http://"""+request.get_host()+"""/log/?email="""+row["email"]+"""' style="width:50px">Terralogic
                                <br>
                                <br> 
                                <a href='http://"""+request.get_host()+"""' target="_blank">http://"""+request.get_host()+"""</a>
                            </div>
                            <div dir="ltr">
                                <br>
                            </div>
                            </div>
                        </div>
                        </div>
                    """
                    # print(body)
                    for i in range(5):
                        try:
                            email=EmailMultiAlternatives('Smart hire from Terralogic',"Exam date dd/mm/yyyy" ,"terralogic.smarthire@gmail.com", to=[row["email"]])
                            email.attach_alternative(body, "text/html")
                            threading.Thread(target=delete_mail, args=(row["email"],)).start()
                            if(str(email.send())=="1"):
                                add_mails("Sent :"+"\t"+datetime.datetime.today().strftime("%Y-%m-%d %H:%M:%S")+"\t"+row["email"])
                                break  
                            else:
                                add_mails("Retrying "+str(i)+" :"+"\t"+datetime.datetime.today().strftime("%Y-%m-%d %H:%M:%S")+"\t"+row["email"])   
                                time.sleep(5)
                        except Exception as e:
                            print(str(e))
                            add_mails("Retrying.. "+str(i)+" :"+"\t"+datetime.datetime.today().strftime("%Y-%m-%d %H:%M:%S")+"\t"+row["email"])
                          
                add_mails("All mail are sent.")
                time.sleep(10)
                check_email_status(request)
                email_sending_status=False
                return HttpResponse("Mail sent successfully",content_type="text")
            except Exception as e:
                print("-------------Main Exce "+str(e))
                email_sending_status=False
                return HttpResponse("fail",content_type="text")
        else:
            return HttpResponse("Invalid user",content_type="text")
    else:
        return HttpResponse("Invalid req",content_type="text")




@csrf_exempt
def sent_cred(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        data=""
        status_val=Email_status.objects.all().values()
        if len(status_val)>0: 
            data=""
            for val in status_val:
                data+="data:"+val['status']+"\n"
            data+="data: \n\n"
        else:
            data="data:  \n\n"
        print(data)
        return HttpResponse(data,content_type='text/event-stream')
    else:
        print("Only for admin\n")
        return HttpResponse("data: Only for admin\n\n",content_type='text')

@csrf_exempt
def cur_email_status(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        data=""
        global email_delivery_status
        if len(email_delivery_status)>0: 
            data=""
            for val in list(set(email_delivery_status)):
                data+="data:"+val+"\n"
        else:
            data="data: Email Status: \n"
        data+="data: \n\n"
        print(data)
        return HttpResponse(data,content_type='text/event-stream')
    else:
        print("Only for admin\n")
        return HttpResponse("data: Only for admin\n\n",content_type='text')

from django.core.files.storage import FileSystemStorage

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))+"/../"
MEDIA_URL = './uploads/'
UPLOAD_DIR = os.path.abspath(os.path.join(BASE_DIR, MEDIA_URL))


def upload(myfile):
    os.system("sudo chown -R :www-data "+UPLOAD_DIR)
    os.system("sudo chmod 777 -R "+UPLOAD_DIR)
    fs = FileSystemStorage(location=UPLOAD_DIR)
    filename = fs.save(myfile.name, myfile)
    os.system("sudo chmod 777 -R "+UPLOAD_DIR)
    return fs.url(filename)

global_temp_update=None

@csrf_exempt
def get_temp_update(request):
    global_temp_update
    if ("admin" in request.session) and (request.session["admin"]!=None):
        if(global_temp_update!=None):
            print(global_temp_update)
            return HttpResponse(global_temp_update,content_type="text")
        else:
            return HttpResponse("No updates",content_type="text")           
    else:
        return HttpResponse("Only admin can",content_type="text")

import xlrd 
from openpyxl import load_workbook

def read_excel(file,use=1):
    data=list()
    if use==0:
        wb = xlrd.open_workbook(BASE_DIR+file) 
        sheet = wb.sheet_by_index(0) 
        sheet.cell_value(0, 0) 
        for i in range(1,sheet.nrows): 
            temp=list()
            for j in range(sheet.ncols): 
                temp.append(sheet.cell_value(i, j))
            data.append(temp)
    else:
        wb = load_workbook(filename=BASE_DIR+file, read_only=True)
        ws = wb['Sheet1']
        for i in range(2,ws.max_row+1):
            temp=list()
            if len(ws[i])!=0:
                for col in ws[i]:
                    if col.value!=None:
                        temp.append(str(col.value))
                if "".join(temp).strip()!="" and len(temp)!=0:
                    data.append(temp)
    return data

import traceback

@csrf_exempt
def bulk_reg(request):
    global global_temp_update
    if ("admin" in request.session) and (request.session["admin"]!=None):
        if request.method == 'POST' and request.FILES['file']:
            if(not request.FILES['file'].name.lower().endswith(('.xls', '.xlsx'))):
                return HttpResponse("Invalid file format",content_type="text")
            uploaded=upload(request.FILES['file'])
            resp=str(uploaded)
            try:
                data=read_excel(uploaded,0)
                # resp+="\n"+str(data)
                if not request.POST._mutable:
                    request.POST._mutable = True
                for val in data:
                    print(len(val))
                    if len(val)>=2 and len(val)<=4:
                        request.POST["name"]=val[0]
                        request.POST["email"]=val[1]
                        request.POST["mob"]=str(round(val[2]))
                        try:
                            if len(val)==4:
                                request.POST["exam"]=Exam.objects.filter(e_name=val[3]).values("id")[0]["id"] 
                            responce=user_signup(request).__dict__
                            resp+="\n"+val[1]+" : "+("".join(responce["_iterator"])).split("&sep;")[1]
                        except Exception as e:
                            print(traceback.format_exc())
                            resp+="\nfor email "+val[1]+" given exam name "+val[3]+" is incorrect"
                        global_temp_update=resp
                    else:
                        resp+="\nInvalid data format"
                # global_temp_update=None
                return HttpResponse( resp,content_type="text")
            except Exception as e:
                print(traceback.format_exc())
                return HttpResponse(resp+"\n"+str(e),content_type="text")
        else:            
            return HttpResponse("Invalid type of Req",content_type="text")
    else:
        return HttpResponse("Only admin can",content_type="text")


@csrf_exempt
def bulk_que(request):
    global global_temp_update
    if ("admin" in request.session) and (request.session["admin"]!=None):
        if request.method == 'POST' and request.FILES['file']:
            if(not request.FILES['file'].name.lower().endswith(('.xls', '.xlsx'))):
                return HttpResponse("Invalid file format",content_type="text")
            uploaded=upload(request.FILES['file'])
            resp=str(uploaded)
            try:
                data=read_excel(uploaded)
                # resp+="\n"+str(data)
                if not request.POST._mutable:
                    request.POST._mutable = True
                for val in data:
                    if len(val)>=7 and len(val)<=8:
                        request.POST["quetions"]=val[0]
                        request.POST["formated_para"]=val[1]
                        if val[1]=="" or val[1]=="-" or val[1].lower()=="null" or val[1].lower()=="none":
                            request.POST["formated_para"]="nothing"
                        request.POST["opt1"]=val[2]
                        request.POST["opt2"]=val[3]
                        request.POST["opt3"]=val[4]
                        request.POST["opt4"]=val[5]
                        request.POST["ans"]=val[int(val[6])+1]
                        if len(val)==8 and val[7]!="":
                            request.POST["cat"]=val[7]
                        responce=addque(request).__dict__
                        que_len=len(val[0])
                        if que_len>100:
                            que_len=100
                        resp+="\n"+val[0][:que_len-1]+" : "+"".join(responce["_iterator"])                    
                    else:
                        print(len(val))
                        print(val)
                        resp+="\nInvalid data format"
                    global_temp_update=resp
                # global_temp_update=None
                return HttpResponse( resp,content_type="text")
            except Exception as e:
                return HttpResponse(resp+"\n"+str(e),content_type="text")
        else:            
            return HttpResponse("Invalid type of Req",content_type="text")
    else:
        return HttpResponse("Only admin can",content_type="text")


@csrf_exempt
def bulk_code_que(request):
    global global_temp_update
    if ("admin" in request.session) and (request.session["admin"]!=None):
        if request.method == 'POST' and request.FILES['file']:
            if(not request.FILES['file'].name.lower().endswith(('.xls', '.xlsx'))):
                return HttpResponse("Invalid file format",content_type="text")
            uploaded=upload(request.FILES['file'])
            resp=str(uploaded)
            try:
                data=read_excel(uploaded)
                # resp+="\n"+str(data)
                if not request.POST._mutable:
                    request.POST._mutable = True
                for val in data:
                    if len(val)>=6 and len(val)<=8:
                        que_len=len(val[0])
                        if que_len>100:
                            que_len=100
                        resp+="\n"+val[0][:que_len-1]
                        request.POST["pblm_stmt"]=val[0]
                        request.POST["code"]=val[1]
                        request.POST["sample_input"]=val[2]
                        request.POST["t_inp_1"]=val[3]
                        request.POST["t_inp_2"]=val[4]
                        request.POST["t_inp_3"]=val[5]
                        request.POST["t_inp_4"]=val[6]
                        if len(val)==8:
                            lang=val[7].lower()
                            if lang=="python":
                                lang="py3"
                            elif lang=="javascript":
                                lang="js"                        
                            request.POST["lang"]=lang
                        for i in range(1,5):
                            try:                            
                                request.POST["exp_out_"+str(i)]=str(escape(run_code(request.POST["lang"],request.POST["code"],"validater",request.POST["t_inp_"+str(i)])))
                            except Exception as e:
                                resp+="\n"+str(e)+"\n"
                        responce=addcode(request).__dict__
                        # resp+="\n"+str(json.dumps(request.POST))                        
                        resp+=" : "+"".join(responce["_iterator"])                       
                    else:
                        resp+="\nInvalid data format"
                    global_temp_update=resp
                # global_temp_update=None
                return HttpResponse( resp,content_type="text")
            except Exception as e:
                return HttpResponse(resp+"\n"+str(e),content_type="text")
        else:            
            return HttpResponse("Invalid type of Req",content_type="text")
    else:
        return HttpResponse("Only admin can",content_type="text")


import subprocess,os

def run_cmd(cmd):
    # print(cmd)
    return subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT).stdout.read().decode()


def run_code(command,code,username,inputs=None,args="",pgm_dir="",pgm_name="pgm"):
    print("command: ",command)
    txt=""
    path="/temp"
    user_home=path+"/"+username
    print("username: "+username)
    print(run_cmd("sudo mkdir "+path))
    print(run_cmd("sudo chmod 777 "+path))
    # print(run_cmd("sudo mount -t tmpfs -o size=100m myramdisk "+path))
    print(run_cmd("sudo mkdir "+user_home))
    print(run_cmd("sudo useradd -r -m "+username+" -d "+user_home))
    print(run_cmd("sudo passwd -l "+username))
    print(run_cmd("sudo mkdir "+user_home+"/"+pgm_dir))
    print(run_cmd("sudo chown "+username+":www-data -R "+user_home))
    print(run_cmd("sudo chmod 771 -R "+user_home))
    # print(run_cmd("sudo ls -l -R "+path+" && sudo tree "+path))
    pre_cmd=""
    if(inputs!=None and len(str(inputs))!=0):
        pre_cmd="printf \""+str(inputs)+"\" | "
    cmd=""
    ext=""
    file_name=user_home+"/"+pgm_dir+pgm_name+"."
    if command=='cpp':
        file_name+="cpp"
        cmd="g++ "+file_name+" -o "+user_home+"/a.out && "+pre_cmd+user_home+"/a.out"+" "+args
    elif command=='java':
        file_name+="java"
        cmd="javac "+file_name+" && cd "+user_home+"/"+pgm_dir+" && "+pre_cmd+" java "+pgm_name+" "+args
    elif command=='py':
        file_name+="py"
        cmd=pre_cmd+"python "+file_name+" "+args
    elif command=='py3':
        file_name+="py"
        cmd=pre_cmd+"python3 "+file_name+" "+args
    elif command=='js':
        file_name+="js"
        cmd=pre_cmd+"node "+file_name+" "+args
    else:
        cmd="echo 'Invalid command: "+command+"'"

    # code=re.sub("\s*public\s+class\s+[a-zA-Z0-9]+\s*{","public class pgm{",code)
    
    f = open(file_name, "w")
    f.write(code)
    f.close()
    # print(run_cmd("sudo chown "+username+":www-data -R "+user_home))
    # print(run_cmd("sudo chmod 771 -R "+user_home))
   
    # cmd="sudo login -f students & "+cmd
    cmd="cd "+user_home+"/"+pgm_dir+" && sudo su "+username+" -c '"+cmd+"'"
    # os.system("sudo chmod 777 "+file_name)           
    print(cmd)
    # print("Code: \n"+code)
    txt="\n"+run_cmd(cmd)
    # print(run_cmd("tree "+path+" && ls -l -R "+user_home))
    # print(txt)
    return txt

@csrf_exempt
def get_expected_output(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        out=escape(run_code(request.POST["cmd"],request.POST["code"],"validater",request.POST["inputs"]))
        return HttpResponse(out,content_type="text")
    else:
        return HttpResponse("Only admin can",content_type="text")

@csrf_exempt
def get_code_count(request): 
    if  request.method == 'POST' and ("admin" in request.session) and (request.session["admin"]!=None):
        try:
            stmt="SELECT COUNT(*) AS code_count FROM smart_code_questions"
            code_count=json.loads(make_query(stmt))[0]
            return HttpResponse(json.dumps(code_count),content_type="text")  
        except Exception as e:
            print(str(e))
            return HttpResponse("warning&sep;Enter valid details",content_type="text")  
    else:
        return HttpResponse("error&sep;You are not allowed for do this operation",content_type="text")

@csrf_exempt
def get_code_que(request):
    if ("logged_in" in request.session):
        today=timezone.now()
        try:
            exam=Exam.objects.filter(id=(Users.objects.filter(email=request.session["logged_in"]).values())[0]["exam_id"],start_date__lte=today).values()[0]
        except Exception as e:
            print("range: "+str(e))
            return HttpResponse("&start;",content_type="text")
        try:
            exam=Exam.objects.filter(id=(Users.objects.filter(email=request.session["logged_in"]).values())[0]["exam_id"],end_date__gte=today).values()[0]
        except Exception as e:
            print("range: "+str(e))
            return HttpResponse("&close;",content_type="text")
        user=Users.objects.filter(email=request.session["logged_in"]).values("id","exam")[0]
        username="smarthire_user_"+str(user["id"])+"_"+str(user["exam"])
        if "que_cnt" not in request.session:
            request.session["que_cnt"]=0
        if "selected_code_questions" not in request.session or ("selected_code_questions"  in request.session and len(request.session["selected_code_questions"])<1):
            cursor = connection.cursor()        
            stmt="SELECT q.id,q.pbm_stmt,q.code,q.lang,q.example_input,q.test_input1,q.test_input2,q.test_input3,q.test_input4,q.expected_output1,q.expected_output2,q.expected_output3,q.expected_output4 FROM smart_code_questions as q INNER JOIN smart_selected_code_questions as sq on q.id=sq.code_questions_id WHERE sq.exam_id=\'"+str(user["exam"])+"\' order by RAND();"
            cursor.execute(stmt)
            request.session["selected_code_questions"]= cursor.fetchall()   
        if request.session["que_cnt"]<len(request.session["selected_code_questions"]):
            cur_que=request.session["selected_code_questions"][request.session["que_cnt"]]
            quetions_to_be_send=""
            quetions_to_be_send+=cur_que[1]+"&sep;"
            quetions_to_be_send+=cur_que[4]+"&sep;"
            quetions_to_be_send+=run_code(cur_que[3],unescape(cur_que[2]),"validater",cur_que[4],"",username+"/")+"&sep;"
            print(Exam.objects.filter(id=user["exam"]).values())
            dur=Exam.objects.filter(id=user["exam"]).values()[0]["code_duration"]
            quetions_to_be_send+=str(dur)
            print(quetions_to_be_send)
            try:
                crs=Coding_result_set()
                crs.user_id=int(user["id"])
                crs.code_questions_id=int(cur_que[0])
                crs.exam_id=int(user["exam"])
                crs.s_time=datetime.datetime.now().strftime("%H:%M:%S")
                crs.e_time=datetime.datetime.now().strftime("%H:%M:%S")
                crs.save()
            except Exception as e:
                print(str(e))
            return HttpResponse(quetions_to_be_send,content_type="text")
        else:
            return HttpResponse("&end;",content_type="text")
       
        return HttpResponse("nothing",content_type="text")
    else:
        return HttpResponse("No user logged in",content_type="text")

@csrf_exempt
def ver_code_que(request):
    txt="Welcome: \n"+request.get_host()+"\n"
    if ("admin" in request.session) and (request.session["admin"]!=None):
        print("admin\t",request.POST["code"])
        txt+="as an Admin\n"
        txt+="\n"+request.POST["code"]
        txt+="\n"+subprocess.Popen(request.POST["code"], shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT).stdout.read().decode()
    elif ("logged_in" in request.session):
        if request.method == 'POST' and  (check_code_attend(request) or "que_cnt" in request.session):
            if re.match("^(c|cpp|java|py|py3|js|cat)$",request.POST["cmd"]):
                # sudo login -f students
                txt+="as an "+request.session["logged_in"]+"\n"
                user=Users.objects.filter(email=request.session["logged_in"]).values("id","exam")[0]
                username="smarthire_user_"+str(user["id"])+"_"+str(user["exam"])
                if request.session["que_cnt"]<len(request.session["selected_code_questions"]):
                    cur_que=request.session["selected_code_questions"][request.session["que_cnt"]]
                    actual_output=""
                    score=0
                    print(request.POST["code"])
                    if("inputs" in request.POST):
                        if (len(cur_que[4])>=1 and len(request.POST["inputs"])>=1) or len(cur_que[4])<1:
                            expected_output=escape(run_code(cur_que[3],unescape(cur_que[2]),"validater",cur_que[4],"",username+"/"))
                            actual_output=escape(run_code(request.POST["cmd"],request.POST["code"],username,cur_que[4],request.POST["arg"]))
                            if expected_output==actual_output:
                                score+=1
                            for i in range(5,9):
                                try:
                                    expected_output=escape(cur_que[i+4])
                                    actual_output=escape(run_code(request.POST["cmd"],request.POST["code"],username,cur_que[i],request.POST["arg"]))               
                                    if expected_output.strip()==actual_output.strip():
                                        score+=1
                                except Exception as e:
                                    print(str(e))   
                            actual_output=run_code(request.POST["cmd"],request.POST["code"],username,request.POST["inputs"],request.POST["arg"])      
                        else:
                            txt+="Input is needed\n"      
                    expected_output=run_code(cur_que[3],cur_que[2],"validater",cur_que[4],"",username+"/")     
                    data={}
                    data["user"]=txt
                    data["qno"]=request.session["que_cnt"]
                    data["output"]=actual_output
                    data["score"]=score
                    txt=json.dumps(data)
                    if "coding_score" not in request.session:
                        request.session["coding_score"]=0
                    request.session["coding_score"]+=score
                    e_time=datetime.datetime.now().strftime("%H:%M:%S")
                    Coding_result_set.objects.filter(user_id=str(user["id"]),code_questions=int(cur_que[0])).update(user_code=escape(request.POST["code"]),lang=request.POST["cmd"],e_time=e_time,total_testcase_passed=score)
                else:
                    txt="&end;"

            else:
                cmd="echo 'invalid cmd'"
        else:
            txt="&closed;"
    else:
        txt+="No user logged in"
    # que = Questions.objects.all().values()
    # for q in que:
    #     txt+=str(q["id"])+" Q:"+str(Questions.objects.filter(id=q["id"]).update(question="<div>"+escape(q["question"][5:len(q["question"])-6])+"</div>"))
    #     txt+="\n"
    #     txt+=str(q["id"])+" opt1:"+str(Questions.objects.filter(id=q["id"]).update(opt1=escape(q["opt1"])))
    #     txt+="\n"
    #     txt+=str(q["id"])+" opt2:"+str(Questions.objects.filter(id=q["id"]).update(opt2=escape(q["opt2"])))
    #     txt+="\n"
    #     txt+=str(q["id"])+" opt3:"+str(Questions.objects.filter(id=q["id"]).update(opt3=escape(q["opt3"])))
    #     txt+="\n"
    #     txt+=str(q["id"])+" opt4:"+str(Questions.objects.filter(id=q["id"]).update(opt4=escape(q["opt4"])))
    #     txt+=str(q["id"])+" ans:"+str(Questions.objects.filter(id=q["id"]).update(ans=escape(q["ans"])))
    #     txt+="\n"
    #     txt+="\n\n"
    # get_inbox_mails("cxg@fg.fygdrt",request)
    return HttpResponse(txt,content_type="text")



@csrf_exempt
def submit_code(request):
    if ("logged_in" in request.session):
        print("----------------next que--------------------")
        if request.session["que_cnt"]<len(request.session["selected_code_questions"]) and "coding_score" in request.session:
            Users.objects.filter(email=request.session["logged_in"]).update(coding_score=str(request.session["coding_score"])+"/"+str(5*len(request.session["selected_code_questions"])))        
        cur_que=request.session["selected_code_questions"][request.session["que_cnt"]]
        e_time=datetime.datetime.now().strftime("%H:%M:%S")
        Coding_result_set.objects.filter(code_questions=int(cur_que[0])).update(e_time=e_time)
        if "que_cnt" not in request.session:
            request.session["que_cnt"]=0
        else:
            request.session["que_cnt"]+=1
        return HttpResponse("next \n",content_type="text")
    else:
        return HttpResponse("No user logged in: \n",content_type="text")
 

def check_code_attend(request):  
    # user = Users.objects.filter(email=request.session["logged_in"],coding_score="-1").values()
    user=Coding_result_set.objects.filter(user_id=(Users.objects.filter(email=request.session["logged_in"]).values())[0]["id"])  
    return not len(user)>0


@csrf_exempt
def code_exam_status(request): 
    if check_code_attend(request):
        return HttpResponse("open",content_type="text")
    else:
        return HttpResponse("closed",content_type="text")

@csrf_exempt
def code_exam_logout(request): 
    if("logged_in" in request.session):
        if("que_cnt" in request.session):
            del request.session['que_cnt']
        return HttpResponse("Deleted",content_type="text")
    else:
        return HttpResponse("closed",content_type="text")

@csrf_exempt
def video_stream(request):
    print("post",request.POST)
    print("get",request.GET)
    return HttpResponse( "GET:\n"+str(request.GET)+"\n\nPOST:\n"+str(request.POST),content_type="text")

@csrf_exempt
def addcode(request):      
    if ("admin" in request.session) and (request.session["admin"]!=None):
        try:
            for key in request.POST:
                try:
                    print("'"+request.POST[key]+"'")
                    if request.POST[key].strip()=="" or request.POST[key]==None:
                        return HttpResponse("Failed",content_type="text")
                except:
                    pass

            code_que = Code_questions()
            code_que.pbm_stmt=request.POST.get("pblm_stmt")
            code_que.code=escape(request.POST.get("code"))
            code_que.lang=request.POST.get("lang")
            code_que.example_input=request.POST.get("sample_input")
            code_que.test_input1=request.POST.get("t_inp_1")
            code_que.test_input2=request.POST.get("t_inp_2")
            code_que.test_input3=request.POST.get("t_inp_3")
            code_que.test_input4=request.POST.get("t_inp_4")
            code_que.expected_output1=request.POST.get("exp_out_1")
            code_que.expected_output2=request.POST.get("exp_out_2")
            code_que.expected_output3=request.POST.get("exp_out_3")
            code_que.expected_output4=request.POST.get("exp_out_4")
            code_que.save()
            if code_que!=None:
                return HttpResponse("success",content_type="text")
            else:
                return HttpResponse("Failed",content_type="text")     
        except Exception as e:
            print(str(e))
            return HttpResponse("Failed",content_type="text")   
    else:
        return HttpResponse("invalid req",content_type="text")


@csrf_exempt
def view_code_res(request):      
    if ("admin" in request.session) and (request.session["admin"]!=None):
        if request.method == 'POST':
            stmt='''
            SELECT 
                cq.pbm_stmt AS 'Problem statement',
                crs.user_code AS 'User code',
                crs.lang 'Programming language',
                crs.total_testcase_passed AS "Number of testcase passed",
                ABS(TIME_TO_SEC(crs.e_time)-TIME_TO_SEC(crs.s_time)) AS 'Duration'''+escape("(Mins)")+''''

            FROM smart_coding_result_set AS crs
                INNER JOIN smart_code_questions AS cq
                    ON cq.id=crs.code_questions_id
                INNER JOIN smart_users su
                    ON su.id=crs.user_id
                WHERE su.id=\''''+request.POST.get("id")+'''\'

            '''
            return HttpResponse(make_query(stmt),content_type="text")
        else:
            return HttpResponse("Invalid req",content_type="text")
    else:
        return HttpResponse("Only for admin",content_type="text")


    
@csrf_exempt
def add_code_que_set(request):  
    if  request.method == 'POST' and ("admin" in request.session) and (request.session["admin"]!=None):
        try:
            print(request.POST.get("exam"),request.POST.get("total"),request.POST.get("dur"))
            try:
                stmt="SELECT abs(TIMESTAMPDIFF(SECOND,se.start_date,se.end_date)) as diff FROM smart_exam as se WHERE id="+str(request.POST.get("exam"))+";"
                max_dur=json.loads(make_query(stmt))[0]["diff"]
                dur=Exam.objects.filter(id=request.POST.get("exam")).values()[0]["duration"]
                print(int(dur)+abs(int(request.POST.get("dur")))*60,max_dur)
                if (int(dur)+abs(int(request.POST.get("dur")))*60)>int(max_dur):
                    return HttpResponse("error&sep;Maximum exam durartion is "+str(int(int(max_dur)/60))+" Mins",content_type="text")
            except:
                pass
            try:
                Selected_code_questions.objects.filter(exam_id=request.POST.get("exam")).delete()
            except Exception as e:
                print(str(e))
            cursor = connection.cursor()     
            stmt="select id from smart_code_questions order by RAND() LIMIT "+str(request.POST.get("total"))+";"
            print(stmt)
            cursor.execute(stmt)
            questions= cursor.fetchall() 
            print(questions)
            for val2 in questions:
                selected_questions=Selected_code_questions()
                selected_questions.code_questions_id=val2[0]
                selected_questions.exam_id=request.POST.get("exam")
                cursor = connection.cursor()
                duration=abs(int(request.POST.get("dur")))*60
                stmt="UPDATE smart_exam SET code_duration = \'"+str(duration)+"\',code_total_marks = \'"+str(len(questions)*5)+"\' WHERE id=\'"+request.POST.get("exam")+"\';"
                cursor.execute(stmt)
                cursor.close()
                try:
                    selected_questions.save()
                except:
                    pass  
            return HttpResponse("success&sep;Exam Populated",content_type="text")  
        except Exception as e:
            return HttpResponse("warning&sep;Enter valid details",content_type="text")  
    else:
        return HttpResponse("Invalid req",content_type="text")  

@csrf_exempt
def list_code_questions(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        stmt='''
        SELECT 
            id,pbm_stmt,code,lang,example_input,test_input1,test_input2,test_input3,test_input4 
        FROM 
            smart_code_questions
        '''
        return HttpResponse(make_query(stmt),content_type="text")
    else:
        return HttpResponse("Only for admin",content_type="text")

@csrf_exempt
def del_code_questions(request):
    if ("admin" in request.session) and (request.session["admin"]!=None):
        int_list=list(map(int,request.POST.get("ids").split(",")))
        print(str(request.POST.get("ids").split(","))+"\n"+str(int_list))
        resp=""
        try:
            resp=str(Code_questions.objects.filter(id__in=int_list ).delete())+"\n"
            return HttpResponse("sucess: "+resp,content_type="text")
        except Exception as e:
            print(str(e))
            return HttpResponse("fail",content_type="text")
    else:
        return HttpResponse("Invalid req",content_type="text")

def get_all_exam_status(uid):
    stmt='''
   
    SELECT 
        t1.total as total_code_que,
        t2.total as total_mcq_que,
        rs1.cnt as attended_code_que,
        rs2.cnt as attended_mcq_que,
        exam.start_date,
        exam.end_date,
        CAST(exam.duration/60 AS DECIMAL(16,0)) as  duration,
        CAST(exam.code_duration/60 AS DECIMAL(16,0)) as  code_duration,
        IF(t1.total>0,IF(rs1.cnt>0,1,0),1) as attended_code,
        IF(t2.total>0,IF(rs2.cnt>0,2,0),2) as attended_mcq,
        IF(t1.total>0,IF(rs1.cnt>0,1,0),1)+IF(t2.total>0,IF(rs2.cnt>0,2,0),2) as status_code
        
    from
        (
            SELECT
                count(*) as total
            from
                smart_selected_code_questions scq 
            where
                scq.exam_id=
                (
                    select 
                        su1.exam_id 
                    from 
                        smart_users su1
                    WHERE 
                        su1.id={0}
                )
        ) as t1
    JOIN
        (
            SELECT
                count(*) as total 
            from
                smart_selected_questions sq 
            where
                sq.exam_id=
                (
                    select 
                        su2.exam_id 
                    from 
                        smart_users su2
                    WHERE 
                        su2.id={0}
                )
        ) as t2
    JOIN
        (
            SELECT
                count(*) as cnt
            from 
                smart_coding_result_set
            WHERE
                smart_coding_result_set.user_id={0}
        ) as rs1
    JOIN
        (
            SELECT
                count(*) as cnt
            from 
                smart_result_set
            WHERE
                smart_result_set.user_id={0}
        ) as rs2
     JOIN
    	(
            SELECT 
            	se.duration,
            	se.code_duration,
            	se.start_date,
            	se.end_date
            FROM
            	smart_exam  se
            WHERE
				se.id=(
                                select 
                                    su2.exam_id 
                                from 
                                    smart_users su2
                                WHERE 
                                    su2.id={0}
                           )
         )  as exam
    
    '''.format(uid)
    # print(stmt)
    return make_query(stmt)

@csrf_exempt
def all_exam_status(request): 
    if("logged_in" in request.session):
        try:
            uid=Users.objects.filter(email=request.session["logged_in"]).values()[0]["id"]
            return HttpResponse(get_all_exam_status(uid),content_type="text")
        except:
            return HttpResponse("error",content_type="text")
    else:
        return HttpResponse("No user logged in: \n",content_type="text")



@csrf_exempt
def log_request(request): 
    email="Unknown"
    if("email" in request.GET):
        email=request.GET["email"]
    #  path = os.getcwd()+"/public/Terralogo.png"
    path = "./public/Terralogo.png"
    ip="Unknown"
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    time=datetime.datetime.now().strftime("%Y-%m-%d-%I-%M-%S-%p")
    print(ip,email,time,request.META['REMOTE_ADDR'],path)
    # return HttpResponse(str(path),content_type="text")
    try:
        email_status=Email_open_status()
        email_status.email=email
        email_status.ip=ip
        email_status.time=time
        email_status.save()
    except Exception as e:
        print(str(e))        
    try:
        return FileResponse(open(path, 'rb'), content_type="image/png")
    except Exception as e:
        print(str(e))
        return HttpResponse("No image",content_type="text")