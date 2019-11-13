from datetime import datetime
import pytz


def utc_to_local(local):
   # METHOD 1: Hardcode zones:
    from_zone = pytz.timezone('UTC')
    to_zone = pytz.timezone('Asia/Kolkata')

    # # METHOD 2: Auto-detect zones:
    # from_zone = tz.tzutc()
    # to_zone = tz.tzlocal()

    # utc = datetime.utcnow()
    utc = datetime.strptime(local, '%Y-%m-%d %H:%M:%S')

    # Tell the datetime object that it's in UTC time zone since 
    # datetime objects are 'naive' by default
    utc = utc.replace(tzinfo=from_zone)

    # Convert time zone
    return utc.astimezone(to_zone).strftime("%Y-%m-%d %H:%M:%S")

print(utc_to_local('2019-09-18 00:00:00'))
# from django.utils import timezone
# date=datetime(2014, 8, 1, 20, 15, 0, 513000, timezone.utc)

# timezone.localtime(date)
# # datetime.datetime(2014, 8, 1, 16, 15, 0, 513000, tzinfo=<DstTzInfo 'America/New_York' EDT-1 day, 20:00:00 DST>)
#2019-09-18T05:30:00+05:30
from openpyxl import load_workbook
wb = load_workbook(filename='uploads/questions.xlsx', read_only=True)
ws = wb['Sheet1']
rows=ws.rows
for i in range(2,ws.max_row):
    temp=""
    if len(ws[i])!=0:
        # if ws[i][0].value!=None:
        print(ws[i])
        for col in ws[i]:
            if col.value!=None:
                temp+=str(col.value)+","
        if "".join(temp).strip()!="":
            print(temp)